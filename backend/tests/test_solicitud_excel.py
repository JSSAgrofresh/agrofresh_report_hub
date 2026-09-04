"""
El Excel de una solicitud: columnas dinámicas por analito+dosis, y que el
lector de "Reporte de Cromatografía" (emitir.py) lo siga entendiendo después
del cambio.

No necesita Postgres: son funciones puras sobre bytes de Excel.
"""
from __future__ import annotations

import io

import openpyxl

from app.emitir import _mapear_solicitud_a_campos
from app.solicitud_excel import (
    _analitos_fungicidas,
    construir_workbook,
    construir_workbook_exportacion,
    leer_datos_workbook,
)
from app.toma_muestras import ANALITOS_DEFECTO

DATOS_BASE = {
    "laboratorio": "AGROFRESH",
    "solicitante": "J",
    "sold_to": "ZZ-TEST",
    "ship_to": None,
    "especie": "Cerezas",
    "variedad": None,
    "linea_proceso": None,
    "csg": None,
    "lote": None,
    "posicion_muestreo": None,
    "numero_camara": None,
    "numero_orden": None,
    "kilos_procesados": None,
    "producto_utilizado": None,
    "tipo_muestra": None,
    "fecha_muestreo": None,
    "hora_muestreo": None,
    "nombre_muestreador": None,
    "generado_por": "J",
    "email_solicitante": None,
    "email_laboratorio": None,
    "observacion": None,
    "numero_solicitud": "OT-0001",
    "fecha_solicitud": "2026-09-01",
    "creado_en": "2026-09-01T10:00:00+00:00",
    "codigo_muestra": None,
    "enviada": False,
    "enviado_en": None,
    "campos_laboratorio": {
        "Fludioxonil (ppm)": "25",
        "Pirimetanil (ppm)": "15",
        "Tebuconazol (ppm)": "Solicitado",  # se pidió pero sin dosis anotada
        "Tipo Aplicación": "Actimist",
    },
    "analitos_solicitados": ["FDL", "PYR", "TEBU"],
}


def _headers_y_fila(ws):
    headers = [ws.cell(row=2, column=c).value for c in range(1, ws.max_column + 1)]
    fila = [ws.cell(row=3, column=c).value for c in range(1, ws.max_column + 1)]
    return dict(zip(headers, fila))


# --- CASO 2: Excel con una columna por analito + su columna de dosis -------


def test_columnas_dinamicas_por_analito_y_dosis():
    wb = construir_workbook_exportacion([DATOS_BASE], ANALITOS_DEFECTO)
    ws = wb.active
    headers = [ws.cell(row=2, column=c).value for c in range(1, ws.max_column + 1)]

    # No hardcodeado: sale del catálogo (ANALITOS_DEFECTO trae 7 fungicidas
    # para QUITECA y otros 7 -mismos códigos- para AGROFRESH, deduplicados).
    codigos_esperados = [a["codigo"] for a in _analitos_fungicidas(ANALITOS_DEFECTO)]
    assert codigos_esperados  # el catálogo trae algo, si no la prueba no prueba nada

    for codigo in codigos_esperados:
        assert codigo in headers
        assert f"{codigo} Dosis" in headers
        # La columna de dosis va INMEDIATAMENTE después de la del código.
        assert headers.index(f"{codigo} Dosis") == headers.index(codigo) + 1


def test_la_dosis_de_cada_analito_es_la_suya_no_una_mezclada():
    wb = construir_workbook_exportacion([DATOS_BASE], ANALITOS_DEFECTO)
    fila = _headers_y_fila(wb.active)

    assert fila["FDL"] == "✓"
    assert fila["FDL Dosis"] == "25"
    assert fila["PYR"] == "✓"
    assert fila["PYR Dosis"] == "15"
    # Solicitado pero sin dosis real anotada: la columna de dosis queda vacía
    # (no aparece la palabra "Solicitado" en una celda de dosis).
    assert fila["TEBU"] == "✓"
    assert fila["TEBU Dosis"] is None
    # No solicitado: ambas columnas vacías -mismo criterio que hoy (vacío,
    # no un cero ni un guion).
    assert fila["AZOX"] is None
    assert fila["AZOX Dosis"] is None


def test_un_analito_no_solicitado_no_muestra_dosis_aunque_el_catalogo_la_tenga():
    """Un analito que el catálogo sabe pero esta solicitud no pidió nunca
    debe mostrar una dosis, aunque `campos_laboratorio` trajera basura para
    esa etiqueta (ej. una solicitud vieja con datos sueltos)."""
    datos = {
        **DATOS_BASE,
        "analitos_solicitados": ["FDL"],
        "campos_laboratorio": {
            **DATOS_BASE["campos_laboratorio"],
            # Pirimetanil tiene un valor guardado, pero ya no está en
            # analitos_solicitados.
        },
    }
    datos["analitos_solicitados"] = ["FDL"]
    wb = construir_workbook_exportacion([datos], ANALITOS_DEFECTO)
    fila = _headers_y_fila(wb.active)
    assert fila["PYR"] is None
    assert fila["PYR Dosis"] is None


def test_el_excel_individual_usa_la_misma_matriz_dinamica():
    wb = construir_workbook(DATOS_BASE, ANALITOS_DEFECTO)
    ws = wb["Solicitudes"]
    fila = _headers_y_fila(ws)
    assert fila["FDL Dosis"] == "25"
    assert fila["PYR Dosis"] == "15"


# --- CASO 8: el lector de Reporte de Cromatografía sigue funcionando -------


def test_leer_datos_workbook_sigue_reconstruyendo_la_solicitud_completa():
    """El lector de emitir.py (leer_solicitudes_de -> leer_datos_workbook)
    lee la hoja oculta "_data", no la hoja visible que acaba de cambiar de
    estructura. Esta prueba es justamente la que habría fallado si el cambio
    del Excel visible hubiera tocado esa hoja oculta."""
    wb = construir_workbook(DATOS_BASE, ANALITOS_DEFECTO)
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    datos_leidos = leer_datos_workbook(buffer)

    assert datos_leidos["numero_solicitud"] == "OT-0001"
    assert datos_leidos["analitos_solicitados"] == ["FDL", "PYR", "TEBU"]
    assert datos_leidos["campos_laboratorio"]["Fludioxonil (ppm)"] == "25"
    assert datos_leidos["campos_laboratorio"]["Pirimetanil (ppm)"] == "15"


def test_emitir_mapea_cada_analito_con_su_dosis_para_cromatografia():
    """`emitir.py` (Reporte de Cromatografía) arma sus "campos" mezclando los
    generales con `campos_laboratorio` tal cual -no parsea la hoja visible-,
    así que cada analito con su dosis individual le sigue llegando bien."""
    wb = construir_workbook(DATOS_BASE, ANALITOS_DEFECTO)
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    datos = leer_datos_workbook(buffer)

    campos = _mapear_solicitud_a_campos(datos)

    assert campos["Fludioxonil (ppm)"] == "25"
    assert campos["Pirimetanil (ppm)"] == "15"


def test_leer_datos_workbook_no_se_confunde_con_las_columnas_nuevas():
    """Round-trip completo: crear -> guardar -> abrir con openpyxl "de
    afuera" (como si alguien lo bajara) -> confirmar que la hoja visible
    nueva no rompe la hoja `_data`."""
    wb = construir_workbook(DATOS_BASE, ANALITOS_DEFECTO)
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    wb_reabierto = openpyxl.load_workbook(buffer, read_only=True, data_only=True)
    assert "_data" in wb_reabierto.sheetnames
    assert "Solicitudes" in wb_reabierto.sheetnames
    ws_visible = wb_reabierto["Solicitudes"]
    headers = [ws_visible.cell(row=2, column=c).value for c in range(1, ws_visible.max_column + 1)]
    assert "FDL Dosis" in headers
    wb_reabierto.close()
