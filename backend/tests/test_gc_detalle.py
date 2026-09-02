"""
La vista de detalle del archivo del GC.

Reproduce lo que hacía un convertidor HTML aparte: pasar el reporte del equipo
a planilla. Se comprueba contra el archivo real de una corrida y contra el
Excel que aquella herramienta generaba, porque el objetivo es que dé lo mismo.
"""
import io
import os

import openpyxl
import pytest

from app import emitir
from app.gc_parser import es_codigo_puro, parsear_cabecera_gc, parsear_gc_txt

ARCHIVO = os.path.join(os.path.dirname(__file__), "datos", "GLPrprtB.txt")

pytestmark = pytest.mark.skipif(
    not os.path.exists(ARCHIVO), reason="falta el reporte de GC de ejemplo"
)


@pytest.fixture(scope="module")
def cabecera():
    return [
        emitir.CampoCabeceraOut(seccion=s, campo=c, valor=v)
        for s, c, v in parsear_cabecera_gc(open(ARCHIVO, "rb").read())
    ]


@pytest.fixture(scope="module")
def muestras():
    crudas = parsear_gc_txt(open(ARCHIVO, "rb").read())
    return [
        emitir.MuestraGCDetalleOut(
            codigo=m.codigo,
            seq_line=m.seq_line,
            fecha_inyeccion=m.fecha_inyeccion,
            es_muestra=es_codigo_puro(m.codigo),
            resultados=[
                emitir.ResultadoAnalitoOut(
                    analito=r.analito, codigo=None, area=r.area, amount=r.amount, rettime=r.rettime
                )
                for r in m.resultados
            ],
        )
        for m in crudas
    ]


def _respuesta(muestras, cabecera=()):
    return emitir.generar_excel_detalle_gc(
        emitir.DetalleGCIn(muestras=muestras, cabecera=list(cabecera))
    )


def _libro(muestras, cabecera=()):
    import asyncio

    resp = _respuesta(muestras, cabecera)
    trozos: list[bytes] = []

    async def leer():
        async for t in resp.body_iterator:
            trozos.append(t)

    asyncio.run(leer())
    return openpyxl.load_workbook(io.BytesIO(b"".join(trozos)))


class TestParseo:
    def test_lee_la_corrida_completa(self, muestras):
        """53 viales: 13 de cliente y 40 entre curvas, blancos y controles.
        El detalle los muestra todos — la curva de calibración es justamente
        lo que se mira para saber si el equipo estaba midiendo bien."""
        assert len(muestras) == 53
        assert sum(1 for m in muestras if m.es_muestra) == 13

    def test_conserva_el_tiempo_de_retencion(self, muestras):
        """Se descartaba al parsear. Sin él, la hoja de datos completos no
        puede reproducir el reporte del equipo."""
        assert muestras[0].resultados[0].rettime == 7.63

    def test_un_valor_medido_llega_entero(self, muestras):
        vial_1 = next(m for m in muestras if m.codigo == "1")
        tebu = next(r for r in vial_1.resultados if r.analito == "TEBUCONAZOLE")
        assert tebu.area == pytest.approx(1.00441)
        assert tebu.amount == pytest.approx(0.0488688)


class TestCabecera:
    """Con qué se midió: instrumento, columna y parámetros de la secuencia.
    Es lo que respalda un resultado si alguien lo cuestiona."""

    def test_lee_los_campos_del_equipo(self, cabecera):
        valores = {c.campo: c.valor for c in cabecera}
        assert valores["Instrument"] == "GC 2"
        assert valores["Column Description"] == "TG-OCP-II"
        assert valores["Operator"] == "SYSTEM"

    def test_junta_los_dos_pares_de_una_misma_linea(self, cabecera):
        """`Model# : 26077-5720   Manufacturer: Thermo` son dos campos en una
        sola línea; leer solo el primero perdía la mitad de la ficha."""
        valores = {c.campo: c.valor for c in cabecera}
        assert valores["Model#"] == "26077-5720"
        assert valores["Manufacturer"] == "Thermo"
        assert valores["Diameter"] == "250.00 µm"
        assert valores["Length"] == "30.0 m"

    def test_no_parte_una_etiqueta_en_la_barra(self, cabecera):
        """`Shutdown Cmd/Macro` se leía como `Macro` a secas."""
        assert {c.campo for c in cabecera} >= {"Shutdown Cmd/Macro"}

    def test_une_un_valor_partido_en_dos_lineas(self, cabecera):
        """La ruta de la secuencia no cabe en una línea y el equipo la corta.
        Separada, el valor quedaba a la mitad."""
        secuencia = next(c.valor for c in cabecera if c.campo == "Sequence")
        assert secuencia.endswith("GCNPD SECUENCIA 28-08-26.S")

    def test_una_ruta_de_windows_no_parece_un_campo(self, cabecera):
        """`C:\\Chem32` tiene dos puntos: si se tomara como etiqueta, el valor
        se cortaría ahí."""
        directorio = next(c.valor for c in cabecera if c.campo == "Data Directory")
        assert directorio.startswith("C:\\Chem32\\1\\Data")

    def test_conserva_el_orden_del_archivo(self, cabecera):
        secciones = [c.seccion for c in cabecera]
        assert secciones == sorted(secciones, key=lambda s: 0 if s == secciones[0] else 1)


class TestExcel:
    def test_tiene_las_tres_hojas(self, muestras, cabecera):
        """La información del equipo va PRIMERA: es lo que se mira para
        respaldar un resultado, no algo escondido al final."""
        assert _libro(muestras, cabecera).sheetnames == [
            emitir.HOJA_CABECERA,
            emitir.HOJA_DETALLE,
            emitir.HOJA_POR_VIAL,
        ]

    def test_la_hoja_del_equipo_trae_los_campos(self, muestras, cabecera):
        """El encabezado va en la fila 6: arriba quedan el logo y el título."""
        ws = _libro(muestras, cabecera)[emitir.HOJA_CABECERA]
        assert [c.value for c in ws["B6:D6"][0]] == ["Sección", "Campo", "Valor"]
        filas = {f[2]: f[3] for f in ws.iter_rows(min_row=7, values_only=True)}
        assert filas["Instrument"] == "GC 2"
        assert filas["Operator"] == "SYSTEM"

    def test_la_primera_hoja_lleva_titulo_y_logo(self, muestras, cabecera):
        ws = _libro(muestras, cabecera)[emitir.HOJA_CABECERA]
        assert ws["C2"].value == emitir.TITULO_EXCEL
        assert ws["C2"].font.bold and ws["C2"].font.size == 18
        assert ws.sheet_view.showGridLines is False
        assert len(ws._images) == 1

    def test_la_seccion_no_se_repite_en_cada_fila(self, muestras, cabecera):
        """Repetirla en las 25 filas tapa el dato que se viene a leer."""
        ws = _libro(muestras, cabecera)[emitir.HOJA_CABECERA]
        secciones = [f[1] for f in ws.iter_rows(min_row=7, values_only=True)]
        assert secciones[0] == "Instrumento y columna"
        assert secciones[1] is None
        assert [s for s in secciones if s] == [
            "Instrumento y columna",
            "Parámetros de la secuencia",
        ]

    def test_las_tres_hojas_son_tablas_de_excel(self, muestras, cabecera):
        """Como tabla se filtra y ordena sin darle formato a mano cada vez."""
        wb = _libro(muestras, cabecera)
        estilos = {
            hoja: [t.tableStyleInfo.name for t in wb[hoja].tables.values()]
            for hoja in wb.sheetnames
        }
        assert estilos == {
            emitir.HOJA_CABECERA: [emitir.ESTILO_TABLA_CABECERA],
            emitir.HOJA_DETALLE: [emitir.ESTILO_TABLA_DATOS],
            emitir.HOJA_POR_VIAL: [emitir.ESTILO_TABLA_DATOS],
        }

    def test_el_nombre_sale_del_data_directory(self, muestras, cabecera):
        """La carpeta de la corrida ya identifica la secuencia; el nombre del
        archivo no tiene por qué inventar otro."""
        cd = _respuesta(muestras, cabecera).headers["Content-Disposition"]
        assert "GCNPD_SECUENCIA_280826_" in cd
        assert cd.endswith('.xlsx"')

    def test_sin_data_directory_cae_a_la_fecha(self, muestras):
        cd = _respuesta(muestras).headers["Content-Disposition"]
        assert "Resultados_GC_" in cd

    def test_datos_completos_calza_con_el_convertidor_viejo(self, muestras):
        """371 filas es exactamente lo que sacaba la herramienta anterior con
        este mismo archivo."""
        ws = _libro(muestras)[emitir.HOJA_DETALLE]
        assert ws.max_row - 1 == 371

    def test_por_vial_lleva_ppm_y_area_juntos(self, muestras):
        """Antes eran dos hojas separadas: leer un vial obligaba a saltar de
        una a otra para comparar su concentración contra su área."""
        ws = _libro(muestras)[emitir.HOJA_POR_VIAL]
        encabezados = [c.value for c in ws[1]]
        assert encabezados[:4] == [
            "Seq Line", "Ubicación de la Muestra", "Vial", "Tipo",
        ]
        assert encabezados[4] == "DIFENILAMINA ppm"
        assert encabezados[5] == "DIFENILAMINA área"
        assert ws.max_row - 1 == 53

    def test_un_vial_medido_queda_bien_ubicado(self, muestras):
        ws = _libro(muestras)[emitir.HOJA_POR_VIAL]
        encabezados = [c.value for c in ws[1]]
        fila = next(f for f in ws.iter_rows(min_row=2, values_only=True) if f[2] == "1")
        assert fila[encabezados.index("TEBUCONAZOLE ppm")] == pytest.approx(0.0488688)
        assert fila[encabezados.index("TEBUCONAZOLE área")] == pytest.approx(1.00441)

    def test_sin_muestras_no_genera_nada(self):
        with pytest.raises(Exception):
            emitir.generar_excel_detalle_gc(emitir.DetalleGCIn(muestras=[]))
