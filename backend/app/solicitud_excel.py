"""
Genera y lee el Excel "maestro" de una solicitud de Toma de muestras.

Cada solicitud se guarda como un único archivo .xlsx (ya no .json): la hoja
"Solicitud" es el documento legible pensado para imprimir/enviar, y una hoja
oculta "_data" guarda el JSON completo de la solicitud en una celda, para
poder reconstruirla sin tener que parsear la hoja "bonita" -así listar/ver/
eliminar siguen funcionando exactamente igual que antes, ahora leyendo desde
Excel en vez de JSON.
"""
import json

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.worksheet.worksheet import Worksheet

VERDE_OSCURO = "3D6B1F"
VERDE_MEDIO = "70AD47"
VERDE_CLARO = "EBF5E1"
GRIS_TEXTO = "6B7280"
GRIS_LINEA = "D9DCE1"
GRIS_FILA = "F7F8F6"

_BORDE_INFERIOR = Border(bottom=Side(style="thin", color=GRIS_LINEA))
_BORDE_COMPLETO = Border(*(Side(style="thin", color=GRIS_LINEA),) * 4)

# Campos generales en el orden en que aparecen en el documento. El conjunto
# de claves es el mismo que expone el modelo `SolicitudIn` de toma_muestras.py.
CAMPOS_GENERALES_ETIQUETAS: list[tuple[str, str]] = [
    ("numero_solicitud", "N° Solicitud"),
    ("fecha_solicitud", "Fecha Solicitud"),
    ("fecha_muestreo", "Fecha Muestreo"),
    ("fecha_informe", "Fecha Informe"),
    ("hora_muestreo", "Hora Muestreo"),
    ("laboratorio", "Laboratorio"),
    ("solicitante", "Solicitante"),
    ("sold_to", "Sold To"),
    ("ship_to", "Ship To"),
    ("especie", "Especie"),
    ("variedad", "Variedad"),
    ("linea_proceso", "Línea Proceso"),
    ("csg", "CSG"),
    ("lote", "Lote"),
    ("posicion_muestreo", "Posición Muestreo"),
    ("numero_camara", "N° Cámara"),
    ("numero_orden", "N° Orden"),
    ("kilos_procesados", "Kilos Procesados (KG)"),
    ("producto_utilizado", "Producto Utilizado"),
    ("tipo_muestra", "Tipo Muestra"),
    ("nombre_muestreador", "Nombre Muestreador"),
    ("generado_por", "Generado Por"),
    ("email_solicitante", "Email Solicitante"),
    ("email_laboratorio", "Email Laboratorio"),
    ("observacion", "Observación"),
]

CAMPOS_ANALISIS_ETIQUETAS = [
    "Dosis Aplicada",
    "Tipo Aplicación",
    "Gasto",
    "Analito Pesticida 1",
    "Resultado Pesticida 1",
    "Analito Pesticida 2",
    "Resultado Pesticida 2",
    "Analito Pesticida 3",
    "Resultado Pesticida 3",
]

# Matriz horizontal oficial del formato de Solicitud de Análisis.  Estas
# columnas son deliberadamente fijas: tanto una solicitud como la descarga
# masiva deben verse iguales al archivo maestro, aunque un laboratorio no
# utilice todos los análisis.
GRUPOS_EXPORTACION: list[tuple[str, list[tuple[str, str, str]]]] = [
    ("GENERAL", [("general", clave, etiqueta) for clave, etiqueta in CAMPOS_GENERALES_ETIQUETAS]),
    ("QUITECA / AGROFRESH — RESIDUOS FUNGICIDAS", [
        *[("analito", codigo, f"{codigo} ppm") for codigo in ("FDL", "IMZ", "PYR", "TEBU", "AZOX", "TBZ", "DPA")],
        ("campo", "Dosis Aplicada", "Dosis Aplicada"),
        ("campo", "Tipo Aplicación", "Tipo Aplicación"),
        ("campo", "Gasto", "Gasto"),
    ]),
    ("DIAGNOFRUIT — PATÓGENOS qPCR", [
        ("analito", "LEV", "Levaduras UFC/mL"),
        ("analito", "BOT", "Botrytis conidia/mL"),
        ("analito", "ALT", "Alternaria conidia/mL"),
        ("analito", "GEO", "Geotrichum esporas/mL"),
        ("analito", "PEN", "Penicillium conidia/mL"),
    ]),
    ("ALS — MICROBIOLOGÍA AGUA (FSMA)", [
        ("analito", "ECOLI100", "E. Coli UFC/100mL"),
        ("analito", "COLIF100", "Coliformes Totales UFC/100mL"),
    ]),
    ("ALS — METALES PESADOS (mg/kg)", [
        ("analito", "PB", "Plomo mg/kg"),
        ("analito", "HG", "Mercurio mg/kg"),
        ("analito", "AS", "Arsénico mg/kg"),
        ("analito", "CD", "Cadmio mg/kg"),
        ("analito", "AL", "Aluminio mg/kg"),
    ]),
    ("ALS — MICROBIOLOGÍA ALIMENTO (UFC/g)", [
        ("analito", "HONGOS", "Hongos UFC/g"),
        ("analito", "LEVG", "Levaduras UFC/g"),
        ("analito", "COLIFG", "Coliformes Totales UFC/g"),
        ("analito", "ECOLIG", "Escherichia coli UFC/g"),
        ("analito", "ENTERO", "Recuento Enterobacterias UFC/g"),
        ("analito", "SALM", "Salmonella 25g (P/A)"),
    ]),
    ("ALS — OTROS", [
        ("analito", "CENIZAS", "Cenizas Insolubles en Ácido (%)"),
        ("analito", "AFLAT", "Aflatoxinas Totales B1+B2+G1+G2 (µg/kg)"),
    ]),
    ("ALS — PESTICIDA PUNTUAL (abierto, hasta 3 analitos)", [
        ("campo", "Analito Pesticida 1", "Analito Pesticida 1"),
        ("campo", "Resultado Pesticida 1", "Resultado Pesticida 1"),
        ("campo", "Analito Pesticida 2", "Analito Pesticida 2"),
        ("campo", "Resultado Pesticida 2", "Resultado Pesticida 2"),
        ("campo", "Analito Pesticida 3", "Analito Pesticida 3"),
        ("campo", "Resultado Pesticida 3", "Resultado Pesticida 3"),
    ]),
]


def _titulo_seccion(ws: Worksheet, fila: int, texto: str, columnas: int = 4) -> int:
    ws.merge_cells(start_row=fila, start_column=1, end_row=fila, end_column=columnas)
    celda = ws.cell(row=fila, column=1, value=texto.upper())
    celda.font = Font(bold=True, size=11, color=VERDE_OSCURO)
    celda.fill = PatternFill("solid", fgColor=VERDE_CLARO)
    celda.alignment = Alignment(vertical="center", indent=1)
    ws.row_dimensions[fila].height = 22
    return fila + 1


def _valor_visible(clave: str, valor):
    if valor in (None, ""):
        return "—"
    if clave == "kilos_procesados":
        return f"{valor} kg"
    return valor


def _escribir_pares(ws: Worksheet, fila: int, pares: list[tuple[str, object]]) -> int:
    """Escribe dos pares etiqueta/valor por fila, como ficha operativa."""
    for indice in range(0, len(pares), 2):
        grupo = pares[indice:indice + 2]
        for bloque, (etiqueta, valor) in enumerate(grupo):
            col_etiqueta = 1 + bloque * 2
            col_valor = col_etiqueta + 1
            c1 = ws.cell(row=fila, column=col_etiqueta, value=etiqueta)
            c2 = ws.cell(row=fila, column=col_valor, value=valor)
            c1.font = Font(bold=True, size=9.5, color=GRIS_TEXTO)
            c1.fill = PatternFill("solid", fgColor=GRIS_FILA)
            c2.font = Font(size=10)
            c1.border = _BORDE_COMPLETO
            c2.border = _BORDE_COMPLETO
            c1.alignment = Alignment(vertical="center", wrap_text=True)
            c2.alignment = Alignment(vertical="center", wrap_text=True)
        if len(grupo) == 1:
            ws.cell(row=fila, column=3).border = _BORDE_COMPLETO
            ws.cell(row=fila, column=4).border = _BORDE_COMPLETO
        ws.row_dimensions[fila].height = 22
        fila += 1
    return fila


def _analitos_del_documento(datos: dict, analitos: list[dict] | None) -> list[dict]:
    laboratorio = datos.get("laboratorio")
    configurados = [
        a for a in (analitos or [])
        if a.get("laboratorio") == laboratorio and a.get("activo", True)
    ]
    configurados.sort(key=lambda a: (a.get("categoria", ""), a.get("orden", 0), a.get("nombre", "")))

    # Una solicitud histórica puede contener un código retirado. Se conserva
    # al final para que nunca se pierda del documento lo que realmente pidió.
    codigos_configurados = {str(a.get("codigo") or "") for a in configurados}
    for codigo in datos.get("analitos_solicitados") or []:
        if codigo not in codigos_configurados:
            configurados.append({"codigo": codigo, "nombre": codigo, "unidad": "", "categoria": ""})
    return configurados


def construir_workbook(datos: dict, analitos: list[dict] | None = None) -> Workbook:
    # El Excel individual usa exactamente la misma tabla horizontal que la
    # exportación masiva; simplemente contiene una sola fila.
    wb = construir_workbook_exportacion([datos], analitos or [])

    # Hoja oculta con el JSON completo de la solicitud -permite reconstruir
    # la solicitud (listar/ver/eliminar) sin parsear la hoja visible.
    ws_datos = wb.create_sheet("_data")
    ws_datos.sheet_state = "hidden"
    ws_datos["A1"] = json.dumps(datos, ensure_ascii=False)

    return wb


def _etiqueta_analito(analito: dict) -> str:
    return f"{analito['nombre']} ({analito['unidad']})" if analito.get("unidad") else analito["nombre"]


def construir_workbook_exportacion(solicitudes: list[dict], analitos: list[dict]) -> Workbook:
    """Genera la matriz oficial horizontal: dos filas de encabezado y una
    fila por solicitud. El Excel individual llama a esta misma función, por
    lo que sólo se diferencia en que contiene una única fila de datos."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Solicitudes"

    columnas = [columna for _, grupo in GRUPOS_EXPORTACION for columna in grupo]
    analitos_por_codigo = {
        str(a.get("codigo") or ""): a for a in analitos if a.get("activo", True)
    }

    # Fila 1: bandas agrupadas como en el formato maestro.
    columna_inicio = 1
    for titulo, grupo in GRUPOS_EXPORTACION:
        columna_fin = columna_inicio + len(grupo) - 1
        ws.merge_cells(start_row=1, start_column=columna_inicio, end_row=1, end_column=columna_fin)
        celda = ws.cell(row=1, column=columna_inicio, value=titulo)
        celda.font = Font(bold=True, size=10, color=VERDE_OSCURO)
        celda.fill = PatternFill("solid", fgColor=VERDE_CLARO)
        celda.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        # Colorear también las celdas combinadas para conservar la banda.
        for col_idx in range(columna_inicio, columna_fin + 1):
            ws.cell(row=1, column=col_idx).fill = PatternFill("solid", fgColor=VERDE_CLARO)
            ws.cell(row=1, column=col_idx).border = _BORDE_COMPLETO
        columna_inicio = columna_fin + 1

    # Fila 2: encabezados filtrables.
    for col_idx, (_, _, etiqueta) in enumerate(columnas, start=1):
        c = ws.cell(row=2, column=col_idx, value=etiqueta)
        c.font = Font(bold=True, size=9, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor=VERDE_OSCURO)
        c.border = _BORDE_COMPLETO
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[1].height = 28
    ws.row_dimensions[2].height = 36

    for fila_idx, datos in enumerate(solicitudes, start=3):
        campos_lab: dict = datos.get("campos_laboratorio") or {}
        solicitados = {str(codigo) for codigo in (datos.get("analitos_solicitados") or [])}
        dosis = []
        for codigo in ("FDL", "IMZ", "PYR", "TEBU", "AZOX", "TBZ", "DPA"):
            if codigo not in solicitados:
                continue
            analito = analitos_por_codigo.get(codigo)
            valor_dosis = campos_lab.get(_etiqueta_analito(analito)) if analito else None
            if valor_dosis and valor_dosis != "Solicitado" and valor_dosis not in dosis:
                dosis.append(str(valor_dosis))
        for col_idx, (tipo, clave, _) in enumerate(columnas, start=1):
            if tipo == "general":
                valor = datos.get(clave)
            elif tipo == "analito":
                valor = "✓" if clave in solicitados else None
            elif clave == "Dosis Aplicada":
                valor = ", ".join(dosis) or None
            else:
                valor = campos_lab.get(clave)
            celda = ws.cell(row=fila_idx, column=col_idx, value=valor if valor not in (None, "") else None)
            celda.border = _BORDE_COMPLETO
            celda.alignment = Alignment(
                horizontal="center" if tipo == "analito" else "left",
                vertical="center",
                wrap_text=True,
            )
            if tipo == "analito" and valor:
                celda.font = Font(bold=True, color=VERDE_OSCURO)
        ws.row_dimensions[fila_idx].height = 24

    total_columnas = len(columnas)
    ws.freeze_panes = "A3"
    ws.auto_filter.ref = f"A2:{ws.cell(row=max(2, ws.max_row), column=total_columnas).coordinate}"
    ws.sheet_view.showGridLines = False
    for col_idx in range(1, total_columnas + 1):
        letra = ws.cell(row=2, column=col_idx).column_letter
        etiqueta = ws.cell(row=2, column=col_idx).value or ""
        ws.column_dimensions[letra].width = min(28, max(13, len(str(etiqueta)) * 0.85))

    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.page_setup.orientation = "landscape"
    ws.print_title_rows = "1:2"
    ws.print_area = f"A1:{ws.cell(row=max(2, ws.max_row), column=total_columnas).coordinate}"

    return wb


def leer_datos_workbook(ruta_o_buffer) -> dict:
    wb = load_workbook(ruta_o_buffer, read_only=True, data_only=True)
    try:
        if "_data" not in wb.sheetnames:
            raise ValueError("El archivo Excel no contiene la hoja de datos de la solicitud (_data).")
        valor = wb["_data"]["A1"].value
    finally:
        wb.close()
    if not valor:
        raise ValueError("El archivo Excel no contiene los datos de la solicitud (_data).")
    return json.loads(valor)
