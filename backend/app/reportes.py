import io
import re
from typing import Any

import openpyxl
import psycopg2.errors
from fastapi import APIRouter, Depends, HTTPException, Query
from fastapi.responses import StreamingResponse
from pydantic import BaseModel

from .auth import Usuario, alcance_de_datos, solo_interno, usuario_actual
from .db import conexion, cursor_dict

router = APIRouter(prefix="/api/reportes", tags=["reportes"])


# ---------------------------------------------------------------------------
# Datos para los gráficos: formato largo, una fila por analito medido.
# ---------------------------------------------------------------------------

DATOS_QUERY = """
    SELECT
        s.id AS solicitud_id,
        s.nro_solicitud,
        s.laboratorio,
        s.fecha_muestreo,
        s.fecha_entrada,
        s.especie,
        s.variedad,
        s.semana_muestreo,
        s.mes,
        s.temporada,
        s.tipo_servicio,
        COALESCE(c.nombre, s.sold_to_raw) AS cliente,
        COALESCE(p.nombre, s.ship_to_raw) AS planta,
        pa.tipo_aplicacion,
        COALESCE(a.codigo, r.analito_raw) AS ingrediente,
        r.valor_num,
        r.valor_texto
    FROM solicitud s
    LEFT JOIN resultado r ON r.solicitud_id = s.id
    LEFT JOIN planta p ON p.id = s.planta_id
    LEFT JOIN cliente c ON c.id = p.cliente_id
    LEFT JOIN analito a ON a.id = r.analito_id
    LEFT JOIN producto_aplicado pa ON pa.solicitud_id = r.solicitud_id AND pa.analito_id = r.analito_id
    WHERE s.vigente
    {filtro_cliente}
    ORDER BY s.fecha_muestreo DESC NULLS LAST, s.id DESC
"""


def _filtro_alcance(
    usuario: Usuario,
    cliente: str | None,
    planta: str | None,
) -> tuple[str, dict[str, str]]:
    """El `AND ...` que acota una consulta a lo que esta sesión puede ver.

    El filtro se arma SIEMPRE acá y nunca a partir del parámetro crudo: para
    una cuenta de cliente, `alcance_de_datos` descarta lo que pidió el
    navegador y devuelve lo que dice su fila. Ese es el único punto donde se
    decide, y por eso los cuatro endpoints que muestran datos pasan por él.
    """
    cliente, planta = alcance_de_datos(usuario, cliente, planta)
    condiciones = []
    params: dict[str, str] = {}
    if cliente:
        condiciones.append("COALESCE(c.nombre, s.sold_to_raw) = %(cliente)s")
        params["cliente"] = cliente
    if planta:
        condiciones.append("COALESCE(p.nombre, s.ship_to_raw) = %(planta)s")
        params["planta"] = planta
    return ("AND " + " AND ".join(condiciones)) if condiciones else "", params


@router.get("/resumen")
def resumen(usuario: Usuario = Depends(usuario_actual)) -> dict[str, Any]:
    """Versión liviana para tarjetas de panel (no trae el detalle largo de /datos)."""
    # Sin este filtro, la cuenta de un cliente veía en su panel el total de
    # solicitudes de TODOS los clientes del sistema.
    filtro, params = _filtro_alcance(usuario, None, None)
    with conexion(escribir=False) as conn:
        with cursor_dict(conn) as cur:
            cur.execute(
                f"""
                SELECT count(*) AS total FROM solicitud s
                LEFT JOIN planta p ON p.id = s.planta_id
                LEFT JOIN cliente c ON c.id = p.cliente_id
                WHERE s.vigente {filtro}
                """,
                params,
            )
            total_solicitudes = cur.fetchone()["total"]
            # "Fecha de ingreso" = fecha_entrada (cuándo entró la muestra al laboratorio),
            # ventana móvil de los últimos 7 días contra la fecha real del servidor —
            # nunca un valor fijo, así siempre refleja la semana en la que se está.
            cur.execute(
                f"""
                SELECT count(*) AS total FROM solicitud s
                LEFT JOIN planta p ON p.id = s.planta_id
                LEFT JOIN cliente c ON c.id = p.cliente_id
                WHERE s.vigente AND s.fecha_entrada >= CURRENT_DATE - INTERVAL '7 days' {filtro}
                """,
                params,
            )
            registros_ultima_semana = cur.fetchone()["total"]
    return {"total_solicitudes": total_solicitudes, "registros_ultima_semana": registros_ultima_semana}


@router.get("/datos")
def datos(
    cliente: str | None = Query(
        None, description="Si se pasa, solo trae los datos de este cliente (portal de cliente)."
    ),
    planta: str | None = Query(
        None,
        description="Si se pasa (junto con cliente), acota además a esta sucursal — cuentas "
        "de cliente creadas por Ship To, en vez de por Sold To completo.",
    ),
    usuario: Usuario = Depends(usuario_actual),
) -> dict[str, Any]:
    # El filtro se aplica siempre en el SQL, nunca en el navegador. Y para una
    # cuenta de cliente sale de su sesión, no de estos parámetros: antes bastaba
    # con editar `?cliente=` en la barra de direcciones para ver a otro cliente.
    filtro_cliente, params = _filtro_alcance(usuario, cliente, planta)
    with conexion(escribir=False) as conn:
        with cursor_dict(conn) as cur:
            cur.execute(DATOS_QUERY.format(filtro_cliente=filtro_cliente), params)
            filas = cur.fetchall()
            # Aparte del join con resultado (que solo trae solicitudes con al menos un
            # analito medido), se cuenta el total real de solicitudes cargadas: esto es
            # lo que se muestra como "Total de registros" en el KPI inicial de Report.
            cur.execute(
                f"""
                SELECT count(*) AS total FROM solicitud s
                LEFT JOIN planta p ON p.id = s.planta_id
                LEFT JOIN cliente c ON c.id = p.cliente_id
                WHERE s.vigente {filtro_cliente}
                """,
                params,
            )
            total_solicitudes = cur.fetchone()["total"]
    return {"filas": filas, "total": len(filas), "total_solicitudes": total_solicitudes}


_ENCABEZADOS_EXCEL_DATOS = [
    ("nro_solicitud", "N° Solicitud"),
    ("laboratorio", "Laboratorio"),
    ("fecha_muestreo", "Fecha Muestreo"),
    ("fecha_entrada", "Fecha Entrada"),
    ("especie", "Especie"),
    ("variedad", "Variedad"),
    ("semana_muestreo", "Semana"),
    ("mes", "Mes"),
    ("temporada", "Temporada"),
    ("tipo_servicio", "Tipo de Servicio"),
    ("cliente", "Cliente (Sold To)"),
    ("planta", "Sucursal (Ship To)"),
    ("tipo_aplicacion", "Tipo Aplicación"),
    ("ingrediente", "Ingrediente"),
    ("valor_num", "Valor"),
    ("valor_texto", "Valor (texto)"),
]

_PAT_NO_SEGURO = re.compile(r"[^A-Za-z0-9._-]+")


def _nombre_archivo_datos(cliente: str | None, planta: str | None) -> str:
    partes = [p for p in (cliente, planta) if p]
    base = _PAT_NO_SEGURO.sub("_", " ".join(partes)).strip("_") or "todos"
    return f"Datos_{base}.xlsx"


@router.get("/datos/excel")
def datos_excel(
    cliente: str | None = Query(None, description="Igual que /datos: acota la descarga a este cliente."),
    planta: str | None = Query(None, description="Igual que /datos: acota además a esta sucursal."),
    usuario: Usuario = Depends(usuario_actual),
) -> StreamingResponse:
    """Descarga en Excel de TODO lo que existe para el cliente/sucursal pedido
    -mismo filtro exacto que ve el portal de cliente en pantalla, nunca más
    ni menos-, para que un cliente pueda llevarse su propio historial."""
    # "Mismo filtro exacto" ahora lo garantiza el código y no la buena fe:
    # pantalla y descarga pasan por la misma función.
    cliente, planta = alcance_de_datos(usuario, cliente, planta)
    filtro_cliente, params = _filtro_alcance(usuario, cliente, planta)
    with conexion(escribir=False) as conn:
        with cursor_dict(conn) as cur:
            cur.execute(DATOS_QUERY.format(filtro_cliente=filtro_cliente), params)
            filas = cur.fetchall()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Datos"
    for col_idx, (_clave, etiqueta) in enumerate(_ENCABEZADOS_EXCEL_DATOS, start=1):
        ws.cell(row=1, column=col_idx, value=etiqueta)
    for fila_idx, fila in enumerate(filas, start=2):
        for col_idx, (clave, _etiqueta) in enumerate(_ENCABEZADOS_EXCEL_DATOS, start=1):
            ws.cell(row=fila_idx, column=col_idx, value=fila.get(clave))

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    nombre = _nombre_archivo_datos(cliente, planta)
    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{nombre}"'},
    )


@router.get("/clientes")
def clientes(_: Usuario = Depends(solo_interno)) -> list[str]:
    """Nombres de cliente que ya tienen datos cargados — mismo criterio (COALESCE)
    que /datos, para que la lista calce exacto con lo que aparece en el filtro de
    Report. Se usa para el selector de cliente al crear un usuario tipo Cliente.

    Cerrado a las cuentas de cliente: la lista completa revela quiénes son
    clientes de AgroFresh, que no es asunto de ninguno de ellos."""
    with conexion(escribir=False) as conn:
        with cursor_dict(conn) as cur:
            cur.execute(
                """
                SELECT DISTINCT COALESCE(c.nombre, s.sold_to_raw) AS cliente
                FROM solicitud s
                LEFT JOIN planta p ON p.id = s.planta_id
                LEFT JOIN cliente c ON c.id = p.cliente_id
                WHERE s.vigente AND COALESCE(c.nombre, s.sold_to_raw) IS NOT NULL
                ORDER BY 1
                """
            )
            return [fila["cliente"] for fila in cur.fetchall()]


# ---------------------------------------------------------------------------
# Catálogo de analitos: consulta para todos, edición solo para administradores
# (el frontend controla quién ve los botones; el backend no valida roles
# porque hoy no hay autenticación real todavía).
# ---------------------------------------------------------------------------


class AnalitoIn(BaseModel):
    codigo: str
    nombre: str
    categoria: str
    laboratorio: str
    unidad: str
    limite_deteccion: str | None = None
    limite_cuantificacion: str | None = None
    matriz: str | None = None
    activo: bool = True
    limite_min: float | None = None
    limite_central: float | None = None
    limite_max: float | None = None


class AnalitoUpdate(BaseModel):
    codigo: str | None = None
    nombre: str | None = None
    categoria: str | None = None
    laboratorio: str | None = None
    unidad: str | None = None
    limite_deteccion: str | None = None
    limite_cuantificacion: str | None = None
    matriz: str | None = None
    activo: bool | None = None
    limite_min: float | None = None
    limite_central: float | None = None
    limite_max: float | None = None


ANALITOS_QUERY = """
    SELECT id, codigo, nombre, categoria, laboratorio, unidad, limite_deteccion,
           limite_cuantificacion, matriz, activo, limite_min, limite_central, limite_max
    FROM analito
    ORDER BY laboratorio, codigo
"""


@router.get("/analitos")
def listar_analitos() -> list[dict[str, Any]]:
    with conexion(escribir=False) as conn:
        with cursor_dict(conn) as cur:
            cur.execute(ANALITOS_QUERY)
            return cur.fetchall()


RETURNING_COLS = (
    "id, codigo, nombre, categoria, laboratorio, unidad, limite_deteccion, "
    "limite_cuantificacion, matriz, activo, limite_min, limite_central, limite_max"
)


@router.post("/analitos")
def crear_analito(payload: AnalitoIn) -> dict[str, Any]:
    datos = payload.model_dump()
    columnas = list(datos.keys())
    placeholders = ", ".join(["%s"] * len(columnas))
    with conexion(escribir=True) as conn:
        with cursor_dict(conn) as cur:
            try:
                cur.execute(
                    f"INSERT INTO analito ({', '.join(columnas)}) VALUES ({placeholders}) RETURNING {RETURNING_COLS}",
                    [datos[c] for c in columnas],
                )
            except psycopg2.errors.UniqueViolation as err:
                raise HTTPException(
                    409, f"Ya existe un analito con el código '{payload.codigo}' en el laboratorio '{payload.laboratorio}'."
                ) from err
            return cur.fetchone()


@router.put("/analitos/{analito_id}")
def actualizar_analito(analito_id: int, payload: AnalitoUpdate) -> dict[str, Any]:
    cambios = payload.model_dump(exclude_unset=True)
    if not cambios:
        raise HTTPException(400, "No se enviaron campos para actualizar.")
    set_clause = ", ".join(f"{campo} = %s" for campo in cambios)
    with conexion(escribir=True) as conn:
        with cursor_dict(conn) as cur:
            try:
                cur.execute(
                    f"UPDATE analito SET {set_clause} WHERE id = %s RETURNING {RETURNING_COLS}",
                    [*cambios.values(), analito_id],
                )
            except psycopg2.errors.UniqueViolation as err:
                raise HTTPException(409, "Ya existe otro analito con ese código y laboratorio.") from err
            fila = cur.fetchone()
            if fila is None:
                raise HTTPException(404, "Analito no encontrado.")
    return fila


@router.delete("/analitos/{analito_id}")
def eliminar_analito(analito_id: int) -> dict[str, Any]:
    with conexion(escribir=True) as conn:
        with cursor_dict(conn) as cur:
            try:
                cur.execute("DELETE FROM analito WHERE id = %s RETURNING id", (analito_id,))
            except psycopg2.errors.ForeignKeyViolation as err:
                raise HTTPException(
                    409,
                    "No se puede eliminar: este analito ya tiene resultados o aplicaciones cargadas en la base. "
                    "Puedes desactivarlo en vez de eliminarlo (edítalo y desmarca 'Activo').",
                ) from err
            fila = cur.fetchone()
            if fila is None:
                raise HTTPException(404, "Analito no encontrado.")
    return {"id": analito_id}


# ---------------------------------------------------------------------------
# Límites por analito, especie y tipo de servicio: un mismo analito puede tener
# distintos límites según la fruta y el tipo de servicio (ej. FDL en Cereza vs.
# FDL en Manzana-Actimist). especie="" / tipo_servicio="" es el comodín "aplica
# a todo".
# ---------------------------------------------------------------------------


class LimiteAnalitoIn(BaseModel):
    analito_id: int
    especie: str = ""
    tipo_servicio: str = ""
    limite_min: float | None = None
    limite_central: float | None = None
    limite_max: float | None = None


class LimiteAnalitoUpdate(BaseModel):
    especie: str | None = None
    tipo_servicio: str | None = None
    limite_min: float | None = None
    limite_central: float | None = None
    limite_max: float | None = None


LIMITES_RETURNING = "id, analito_id, especie, tipo_servicio, limite_min, limite_central, limite_max"


@router.get("/limites")
def listar_limites() -> list[dict[str, Any]]:
    with conexion(escribir=False) as conn:
        with cursor_dict(conn) as cur:
            cur.execute(
                f"SELECT {LIMITES_RETURNING} FROM analito_limite ORDER BY analito_id, tipo_servicio, especie"
            )
            return cur.fetchall()


@router.post("/limites")
def crear_limite(payload: LimiteAnalitoIn) -> dict[str, Any]:
    with conexion(escribir=True) as conn:
        with cursor_dict(conn) as cur:
            try:
                cur.execute(
                    f"""
                    INSERT INTO analito_limite (analito_id, especie, tipo_servicio, limite_min, limite_central, limite_max)
                    VALUES (%(analito_id)s, %(especie)s, %(tipo_servicio)s, %(limite_min)s, %(limite_central)s, %(limite_max)s)
                    ON CONFLICT (analito_id, especie, tipo_servicio)
                    DO UPDATE SET limite_min = EXCLUDED.limite_min, limite_central = EXCLUDED.limite_central,
                                  limite_max = EXCLUDED.limite_max, actualizado_en = now()
                    RETURNING {LIMITES_RETURNING}
                    """,
                    payload.model_dump(),
                )
            except psycopg2.errors.ForeignKeyViolation as err:
                raise HTTPException(404, "Analito no encontrado.") from err
            return cur.fetchone()


@router.put("/limites/{limite_id}")
def actualizar_limite(limite_id: int, payload: LimiteAnalitoUpdate) -> dict[str, Any]:
    cambios = payload.model_dump(exclude_unset=True)
    if not cambios:
        raise HTTPException(400, "No se enviaron campos para actualizar.")
    set_clause = ", ".join(f"{campo} = %s" for campo in cambios) + ", actualizado_en = now()"
    with conexion(escribir=True) as conn:
        with cursor_dict(conn) as cur:
            try:
                cur.execute(
                    f"UPDATE analito_limite SET {set_clause} WHERE id = %s RETURNING {LIMITES_RETURNING}",
                    [*cambios.values(), limite_id],
                )
            except psycopg2.errors.UniqueViolation as err:
                raise HTTPException(
                    409, "Ya existe un límite para esa especie y tipo de servicio en este analito."
                ) from err
            fila = cur.fetchone()
            if fila is None:
                raise HTTPException(404, "Límite no encontrado.")
            return fila


@router.delete("/limites/{limite_id}")
def eliminar_limite(limite_id: int) -> dict[str, Any]:
    with conexion(escribir=True) as conn:
        with cursor_dict(conn) as cur:
            cur.execute("DELETE FROM analito_limite WHERE id = %s RETURNING id", (limite_id,))
            fila = cur.fetchone()
            if fila is None:
                raise HTTPException(404, "Límite no encontrado.")
            return {"id": limite_id}
