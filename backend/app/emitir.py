import io
import logging
import os
import re
import zipfile
from datetime import date, datetime
from zoneinfo import ZoneInfo, ZoneInfoNotFoundError

import openpyxl
import psycopg2.errors
from fastapi import APIRouter, File, HTTPException, UploadFile
from fastapi.responses import StreamingResponse
from pydantic import BaseModel

from . import r2
from .db import conexion, cursor_dict
from .gc_parser import (
    NOMBRE_GC_A_CODIGO,
    es_codigo_puro,
    parsear_cabecera_gc,
    parsear_gc_txt,
)
from .informe_pdf import generar_informe_pdf
from .mapeo import LABORATORIO_CATALOGO, calcular_semana
from .solicitud_excel import CAMPOS_GENERALES_ETIQUETAS
from .solicitud_parser import parsear_solicitudes_html
from .storage import _carpeta_raiz as _carpeta_raiz_storage, _nombre_seguro
from .toma_muestras import carpeta_de_cliente, leer_solicitudes_de

def _zona_laboratorio() -> ZoneInfo | None:
    """La zona del laboratorio, o None para usar la del sistema.

    El laboratorio está en Rancagua. La base guarda los instantes en UTC, que
    es lo correcto, pero la fecha que el operador espera ver es la de acá: una
    muestra recibida a las 21:00 no se recibió mañana.

    En Linux las zonas salen del sistema operativo; en Windows hay que
    instalar el paquete `tzdata` (está en requirements.txt). Si falta, esto
    devuelve None y `astimezone(None)` convierte a la hora local del equipo
    -que en este servidor es la misma de Rancagua-. Es un dato de fecha en un
    listado: no justifica que el backend entero no arranque.
    """
    try:
        return ZoneInfo("America/Santiago")
    except ZoneInfoNotFoundError:
        logging.getLogger(__name__).warning(
            "Sin base de zonas horarias (falta el paquete tzdata): la fecha de "
            "recepción usará la hora local del servidor. Instálalo con "
            "'pip install tzdata' para no depender del reloj del equipo."
        )
        return None


ZONA_LABORATORIO = _zona_laboratorio()

_PAT_CODIGO_COLUMNA = re.compile(r"\(([A-Za-z]+)\)\s*$")
_PREFIJO_RESULTADO = "Resultado:"

router = APIRouter(prefix="/api/emitir/cromatografia", tags=["emitir"])

# Fuente principal: solicitudes de Toma de muestras del laboratorio AGROFRESH
# (ver toma_muestras.py) -carpeta STORAGE_DIR/solicitudes/AGROFRESH-.
LABORATORIO_SOLICITUDES = "AGROFRESH"

# Fuente legada: archivos HTML-como-.xls subidos manualmente a Storage antes
# de que existiera el módulo Toma de muestras. Se sigue leyendo para no
# romper solicitudes que ya estén ahí, pero ya no es la fuente principal.
#
# Este texto es el nombre real de una carpeta en Storage, no una etiqueta:
# renombrarlo a "Solicitud de Análisis" dejaría de encontrar lo ya guardado.
CARPETA_SOLICITUDES = "Solicitud de Muestreo"


def _asignar_folios(cur, cantidad: int) -> list[str]:
    """Reserva `cantidad` identificadores de informe consecutivos
    AGF{año}-{n} (ej. AGF2026-1), de forma atómica (INSERT+UPDATE dentro de
    la misma transacción de escritura). El correlativo es por año: nunca se
    reinicia dentro del mismo año, solo al cambiar de año."""
    anio = date.today().year
    cur.execute("INSERT INTO informe_folio_anual (anio, siguiente) VALUES (%s, 1) ON CONFLICT (anio) DO NOTHING", (anio,))
    cur.execute(
        "UPDATE informe_folio_anual SET siguiente = siguiente + %s WHERE anio = %s RETURNING siguiente",
        (cantidad, anio),
    )
    siguiente_tras = cur.fetchone()["siguiente"]
    primero = siguiente_tras - cantidad
    return [f"AGF{anio}-{n}" for n in range(primero, siguiente_tras)]


class InformeConfigOut(BaseModel):
    analizado_por_nombre: str
    analizado_por_cargo: str
    aprobado_por_nombre: str
    aprobado_por_cargo: str
    # Apagado, el informe sale con una sola firma -la de aprobación- abajo a
    # la derecha, en vez de dejar media hoja con una raya y un guión.
    incluir_analista: bool = True


_COLUMNAS_CONFIG = (
    "analizado_por_nombre",
    "analizado_por_cargo",
    "aprobado_por_nombre",
    "aprobado_por_cargo",
)


def leer_config_informe() -> dict:
    """La fila de configuración del informe, aunque falte la migración 0022.

    Actualizar son dos pasos y en ese orden: `git pull` + reiniciar, y después
    correr la migración. En esa ventana la columna `incluir_analista` todavía
    no existe, y sin esto generar un informe respondía 500 con un
    UndefinedColumn que no le dice nada a nadie.

    Postgres aborta la transacción al fallar la consulta, así que el segundo
    intento necesita una conexión nueva; no basta con reintentar en la misma.
    """
    columnas = ", ".join(_COLUMNAS_CONFIG)
    try:
        with conexion(escribir=False) as conn, cursor_dict(conn) as cur:
            cur.execute(f"SELECT {columnas}, incluir_analista FROM informe_config WHERE id = 1")
            return dict(cur.fetchone() or {})
    except psycopg2.errors.UndefinedColumn:
        pass

    with conexion(escribir=False) as conn, cursor_dict(conn) as cur:
        cur.execute(f"SELECT {columnas} FROM informe_config WHERE id = 1")
        fila = cur.fetchone()
    # Sin la columna, el informe sale como salía antes: con las dos firmas.
    return {**dict(fila or {}), "incluir_analista": True}


@router.get("/config-informe")
def obtener_config_informe() -> InformeConfigOut:
    fila = leer_config_informe()
    if not fila:
        raise HTTPException(500, "No existe la configuración del informe (falta la migración 0008).")
    return InformeConfigOut(**fila)


@router.put("/config-informe")
def guardar_config_informe(body: InformeConfigOut) -> InformeConfigOut:
    firmas = (
        body.analizado_por_nombre.strip(),
        body.analizado_por_cargo.strip(),
        body.aprobado_por_nombre.strip(),
        body.aprobado_por_cargo.strip(),
    )
    columnas = ", ".join(_COLUMNAS_CONFIG)
    asignaciones = ", ".join(f"{c} = %s" for c in _COLUMNAS_CONFIG)
    try:
        with conexion() as conn, cursor_dict(conn) as cur:
            cur.execute(
                f"""
                UPDATE informe_config
                SET {asignaciones}, incluir_analista = %s, actualizado_en = now()
                WHERE id = 1
                RETURNING {columnas}, incluir_analista
                """,
                (*firmas, body.incluir_analista),
            )
            return InformeConfigOut(**cur.fetchone())
    except psycopg2.errors.UndefinedColumn:
        pass

    # Falta la 0022: se guardan las firmas igual y el check se ignora, en vez
    # de perder también lo que el usuario acaba de escribir.
    with conexion() as conn, cursor_dict(conn) as cur:
        cur.execute(
            f"""
            UPDATE informe_config
            SET {asignaciones}, actualizado_en = now()
            WHERE id = 1
            RETURNING {columnas}
            """,
            firmas,
        )
        fila = cur.fetchone()
    return InformeConfigOut(**fila, incluir_analista=True)


class ResultadoAnalitoOut(BaseModel):
    analito: str
    codigo: str | None
    area: float | None
    amount: float | None
    # Solo para la vista de detalle; el cruce y el informe no lo usan.
    rettime: float | None = None


class MuestraGCOut(BaseModel):
    codigo: str
    seq_line: int | None
    fecha_inyeccion: str | None
    resultados: list[ResultadoAnalitoOut]


@router.post("/parsear-gc")
async def parsear_gc(archivo: UploadFile = File(...)) -> list[MuestraGCOut]:
    contenido = await archivo.read()
    try:
        muestras = parsear_gc_txt(contenido)
    except ValueError as e:
        raise HTTPException(400, str(e)) from e
    if not muestras:
        raise HTTPException(400, "No se encontró ninguna muestra en el archivo. ¿Es el reporte del GC correcto?")

    # Solo los viales con código puro (ej. GCNPD9826) son muestras reales de
    # cliente cruzables: se descartan curvas de calibración, blancos y
    # controles de limpieza (ej. "GCNPD9775 LIMPIEZA NORMAL MET 2").
    muestras_reales = [m for m in muestras if es_codigo_puro(m.codigo)]
    if not muestras_reales:
        raise HTTPException(
            400,
            "El archivo se leyó bien pero ninguna muestra tiene un código puro "
            "(ej. GCNPD9826) — solo se encontraron curvas, blancos o controles.",
        )

    return [
        MuestraGCOut(
            codigo=m.codigo,
            seq_line=m.seq_line,
            fecha_inyeccion=m.fecha_inyeccion,
            resultados=[
                ResultadoAnalitoOut(
                    analito=r.analito,
                    codigo=NOMBRE_GC_A_CODIGO.get(r.analito),
                    area=r.area,
                    amount=r.amount,
                    rettime=r.rettime,
                )
                for r in m.resultados
            ],
        )
        for m in muestras_reales
    ]


# ---------------------------------------------------------------------------
# Vista de detalle del archivo del GC
#
# Reproduce lo que antes hacía una herramienta HTML aparte: pasar el reporte
# del equipo a planilla. No toca el cruce ni el informe -es solo otra forma de
# mirar el mismo archivo ya cargado-, así que recibe las muestras ya parseadas
# y no vuelve a leer el .txt.
#
# Aquella herramienta sacaba dos hojas pivote, una de área y otra de ppm. Acá
# van juntas: leer un vial obligaba a saltar de hoja en hoja para comparar su
# concentración contra su área, que es justo lo que se hace al revisar.
# ---------------------------------------------------------------------------

HOJA_CABECERA = "Información del GC"
HOJA_DETALLE = "Datos completos"
HOJA_POR_VIAL = "Área y PPM por vial"


class CampoCabeceraOut(BaseModel):
    seccion: str
    campo: str
    valor: str


class DetalleGCOut(BaseModel):
    """Todo lo que la vista de detalle necesita del archivo, de una sola vez."""

    cabecera: list[CampoCabeceraOut]
    muestras: list["MuestraGCDetalleOut"]


class MuestraGCDetalleOut(MuestraGCOut):
    """Una corrida del GC incluye, además de las muestras de cliente, la curva
    de calibración, blancos y controles de limpieza. El cruce los descarta -no
    son de nadie- pero al revisar la corrida son justamente lo que se mira
    para saber si el equipo estaba midiendo bien."""

    es_muestra: bool


@router.post("/parsear-gc/completo")
async def parsear_gc_completo(archivo: UploadFile = File(...)) -> DetalleGCOut:
    """El archivo del GC entero, para la vista de detalle.

    Existe aparte de /parsear-gc a propósito: ese devuelve solo las muestras
    cruzables, y meter acá las curvas y los blancos haría que el escáner de
    viales pudiera "encontrar" un blanco. Son dos preguntas distintas sobre el
    mismo archivo, y conviene que sigan siéndolo.
    """
    contenido = await archivo.read()
    try:
        muestras = parsear_gc_txt(contenido)
    except ValueError as e:
        raise HTTPException(400, str(e)) from e
    if not muestras:
        raise HTTPException(400, "No se encontró ninguna muestra en el archivo. ¿Es el reporte del GC correcto?")

    return DetalleGCOut(
        cabecera=[
            CampoCabeceraOut(seccion=s, campo=c, valor=v)
            for s, c, v in parsear_cabecera_gc(contenido)
        ],
        muestras=[
        MuestraGCDetalleOut(
            codigo=m.codigo,
            seq_line=m.seq_line,
            fecha_inyeccion=m.fecha_inyeccion,
            es_muestra=es_codigo_puro(m.codigo),
            resultados=[
                ResultadoAnalitoOut(
                    analito=r.analito,
                    codigo=NOMBRE_GC_A_CODIGO.get(r.analito),
                    area=r.area,
                    amount=r.amount,
                    rettime=r.rettime,
                )
                for r in m.resultados
            ],
        )
        for m in muestras
        ],
    )


class DetalleGCIn(BaseModel):
    """Lo que devolvió /parsear-gc/completo, tal cual."""

    cabecera: list[CampoCabeceraOut] = []
    muestras: list[MuestraGCDetalleOut]


def _compuestos_en_orden(muestras: list[MuestraGCOut]) -> list[str]:
    """Los compuestos en el orden en que aparecen en el reporte, que es el
    orden del método del equipo — no alfabético, que a nadie le sirve."""
    vistos: list[str] = []
    for m in muestras:
        for r in m.resultados:
            if r.analito not in vistos:
                vistos.append(r.analito)
    return vistos


@router.post("/detalle-gc/excel")
def generar_excel_detalle_gc(body: DetalleGCIn) -> StreamingResponse:
    if not body.muestras:
        raise HTTPException(400, "No hay muestras para exportar.")

    compuestos = _compuestos_en_orden(body.muestras)
    wb = openpyxl.Workbook()

    # ── Hoja 1: con qué se midió. Es lo que respalda un resultado si alguien
    # lo cuestiona, así que va primero y no escondida al final.
    ws0 = wb.active
    ws0.title = HOJA_CABECERA
    for col, texto in enumerate(("Sección", "Campo", "Valor"), start=1):
        ws0.cell(row=1, column=col, value=texto)
    for fila_idx, campo in enumerate(body.cabecera, start=2):
        ws0.cell(row=fila_idx, column=1, value=campo.seccion)
        ws0.cell(row=fila_idx, column=2, value=campo.campo)
        ws0.cell(row=fila_idx, column=3, value=campo.valor)
    ws0.freeze_panes = "A2"
    for col, ancho in zip("ABC", (26, 46, 92)):
        ws0.column_dimensions[col].width = ancho

    # ── Hoja 2: una fila por compuesto de cada vial, como sale del equipo ──
    ws = wb.create_sheet(HOJA_DETALLE)
    encabezados = [
        "Vial", "Tipo", "Seq Line", "Fecha Inyección",
        "RetTime (min)", "Área (pA*s)", "Amount (ppm)", "Compuesto",
    ]
    for col, texto in enumerate(encabezados, start=1):
        ws.cell(row=1, column=col, value=texto)
    fila = 2
    for m in body.muestras:
        for r in m.resultados:
            for col, valor in enumerate(
                [
                    m.codigo, "Muestra" if m.es_muestra else "Control",
                    m.seq_line, m.fecha_inyeccion, r.rettime, r.area, r.amount, r.analito,
                ],
                start=1,
            ):
                ws.cell(row=fila, column=col, value=valor)
            fila += 1
    ws.freeze_panes = "A2"
    for col, ancho in zip("ABCDEFGH", (16, 10, 9, 22, 13, 14, 14, 18)):
        ws.column_dimensions[col].width = ancho

    # ── Hoja 2: un vial por fila, con ppm y área de cada compuesto pegados ──
    ws2 = wb.create_sheet(HOJA_POR_VIAL)
    ws2.cell(row=1, column=1, value="Seq Line")
    ws2.cell(row=1, column=2, value="Vial")
    ws2.cell(row=1, column=3, value="Tipo")
    for i, compuesto in enumerate(compuestos):
        ws2.cell(row=1, column=4 + i * 2, value=f"{compuesto} ppm")
        ws2.cell(row=1, column=5 + i * 2, value=f"{compuesto} área")
    for fila_idx, m in enumerate(body.muestras, start=2):
        ws2.cell(row=fila_idx, column=1, value=m.seq_line)
        ws2.cell(row=fila_idx, column=2, value=m.codigo)
        ws2.cell(row=fila_idx, column=3, value="Muestra" if m.es_muestra else "Control")
        por_analito = {r.analito: r for r in m.resultados}
        for i, compuesto in enumerate(compuestos):
            r = por_analito.get(compuesto)
            ws2.cell(row=fila_idx, column=4 + i * 2, value=r.amount if r else None)
            ws2.cell(row=fila_idx, column=5 + i * 2, value=r.area if r else None)
    ws2.freeze_panes = "D2"
    ws2.column_dimensions["A"].width = 9
    ws2.column_dimensions["B"].width = 18
    ws2.column_dimensions["C"].width = 10
    for col in range(4, 4 + len(compuestos) * 2):
        ws2.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 17

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    nombre = f"Resultados_GC_{datetime.now().strftime('%Y%m%d')}.xlsx"
    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{nombre}"'},
    )


class SolicitudOut(BaseModel):
    archivo: str
    campos: dict[str, str]
    analitos_solicitados: list[str]
    codigo_muestra: str | None = None
    # Cuándo llegó la muestra al laboratorio. No se pregunta en ningún
    # formulario: se llena solo con el instante del cruce.
    fecha_recepcion: str | None = None
    hora_recepcion: str | None = None


def _partir_recepcion(valor: str | None) -> tuple[str | None, str | None]:
    """El instante del cruce, partido en el día y la hora que ve el operador.

    Llega como timestamp con zona ("2026-09-01T18:03:41.779034+00:00"). Se
    convierte a la hora local del laboratorio antes de partirlo: si no, una
    muestra recibida a las 21:00 en Rancagua aparecería recibida al día
    siguiente.
    """
    if not valor:
        return None, None
    try:
        momento = datetime.fromisoformat(valor)
    except ValueError:
        return None, None
    if momento.tzinfo is not None:
        momento = momento.astimezone(ZONA_LABORATORIO)
    return momento.strftime("%Y-%m-%d"), momento.strftime("%H:%M")


def _fecha_iso_a_ddmmyyyy(valor: str | None) -> str | None:
    if not valor:
        return None
    try:
        return datetime.strptime(valor, "%Y-%m-%d").strftime("%d-%m-%Y")
    except ValueError:
        return valor  # ya venía en otro formato (ej. legado); se deja tal cual


def _mapear_solicitud_a_campos(datos: dict) -> dict[str, str]:
    """Convierte una solicitud de Toma de muestras (dict estructurado, leído
    de la hoja oculta "_data" del Excel — ver solicitud_excel.py) al mismo
    formato "campos: dict[etiqueta, valor]" que ya consume todo el resto del
    pipeline de cromatografía (cruce, exportar Excel/PDF, subir a BD). Así el
    parser es estructural (por clave configurada), no depende de una
    plantilla visual, y el pipeline existente no necesita cambios.

    Todos los campos generales configurados se incluyen siempre, aunque
    estén vacíos (para que el informe final los muestre igual), excepto los
    analitos de laboratorio que el usuario no marcó como solicitados -esa
    regla (nunca mostrar un analito no pedido) se mantiene intacta."""
    campos: dict[str, str] = {}
    for clave, etiqueta in CAMPOS_GENERALES_ETIQUETAS:
        valor = datos.get(clave)
        if clave in ("fecha_solicitud", "fecha_muestreo"):
            valor = _fecha_iso_a_ddmmyyyy(valor)
        campos[etiqueta] = str(valor) if valor not in (None, "") else ""

    # Alias legado: el resto del pipeline (subir-bd, informe PDF) espera
    # estas claves exactas, heredadas del formato HTML-como-.xls original.
    campos["Sold To (Nombre)"] = campos.get("Sold To", "")
    campos["Ship To (Nombre)"] = campos.get("Ship To", "")
    # La observación de la solicitud no forma parte de CAMPOS_GENERALES_ETIQUETAS
    # (en el Excel de la solicitud va en su propia sección) pero el informe de
    # análisis sí la necesita como campo independiente, separado del tratamiento.
    campos["Observación"] = str(datos.get("observacion") or "")

    # Campos propios del laboratorio (analitos solicitados, dosis, tipo de
    # aplicación, etc.) — ya vienen con las etiquetas humanas como clave.
    campos.update(datos.get("campos_laboratorio") or {})
    return campos


@router.get("/solicitudes")
def listar_solicitudes() -> list[SolicitudOut]:
    salida: list[SolicitudOut] = []

    # R2 o disco según cómo esté levantado el sistema: lo resuelve
    # `leer_solicitudes_de`, no este módulo.
    for nombre, datos in leer_solicitudes_de(LABORATORIO_SOLICITUDES):
        fecha_recepcion, hora_recepcion = _partir_recepcion(datos.get("recepcion_en"))
        salida.append(
            SolicitudOut(
                archivo=nombre,
                campos=_mapear_solicitud_a_campos(datos),
                analitos_solicitados=datos.get("analitos_solicitados") or [],
                codigo_muestra=datos.get("codigo_muestra"),
                fecha_recepcion=fecha_recepcion,
                hora_recepcion=hora_recepcion,
            )
        )

    carpeta_legado = os.path.join(_carpeta_raiz_storage(), CARPETA_SOLICITUDES)
    if os.path.isdir(carpeta_legado):
        for nombre in sorted(os.listdir(carpeta_legado)):
            ruta = os.path.join(carpeta_legado, _nombre_seguro(nombre))
            if not os.path.isfile(ruta):
                continue
            try:
                with open(ruta, encoding="utf-8-sig") as f:
                    contenido = f.read()
                solicitudes = parsear_solicitudes_html(contenido)
            except (UnicodeDecodeError, ValueError):
                continue
            for s in solicitudes:
                salida.append(SolicitudOut(archivo=nombre, campos=s.campos, analitos_solicitados=s.analitos_solicitados))

    if not salida:
        raise HTTPException(
            404,
            f'Todavía no hay solicitudes de "{LABORATORIO_SOLICITUDES}" — créalas desde Toma de muestras → Nueva solicitud.',
        )
    return salida


class FilaCruceIn(BaseModel):
    """Una solicitud ya cruzada con su vial del GC: sus campos originales tal
    cual (mismas columnas que el archivo de Storage), qué analitos pidió, y el
    resultado (amount/ppm) por código de analito detectado en el vial
    asignado."""

    campos: dict[str, str]
    analitos_solicitados: list[str]
    resultados_por_codigo: dict[str, float | None]
    codigo_vial: str | None = None
    fecha_inyeccion: str | None = None
    # Fecha en que la muestra física llegó al laboratorio: no viene en la
    # solicitud ni en el resultado del GC, se elige a mano en la zona de
    # cruce (formato ISO "YYYY-MM-DD", el que entrega un <input type=date>).
    fecha_recepcion: str | None = None


@router.post("/excel")
def generar_excel(filas: list[FilaCruceIn]) -> StreamingResponse:
    if not filas:
        raise HTTPException(400, "No hay filas para exportar.")

    with conexion() as conn, cursor_dict(conn) as cur:
        folios = _asignar_folios(cur, len(filas))

    # Mismas columnas que el archivo de solicitud original, en el mismo orden
    # (unión por si alguna solicitud trae un campo que otra no tiene), más el
    # folio interno al inicio.
    columnas: list[str] = []
    for fila in filas:
        for columna in fila.campos:
            if columna not in columnas:
                columnas.append(columna)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Solicitudes con resultado"

    ws.cell(row=1, column=1, value="N° Informe")
    ws.cell(row=1, column=2, value="Fecha Recepción")
    for col_idx, columna in enumerate(columnas, start=3):
        ws.cell(row=1, column=col_idx, value=columna)

    for fila_idx, (fila, folio) in enumerate(zip(filas, folios), start=2):
        ws.cell(row=fila_idx, column=1, value=folio)
        ws.cell(row=fila_idx, column=2, value=fila.fecha_recepcion or None)
        for col_idx, columna in enumerate(columnas, start=3):
            valor: str | float | None = fila.campos.get(columna, "") or None
            if columna.startswith(_PREFIJO_RESULTADO):
                # Nunca se escribe el resultado de un analito que esta
                # solicitud no pidió, aunque el vial asignado sí lo haya
                # detectado -es la regla explícita: solicitud y resultado
                # siempre tienen los mismos analitos, sin excepción-.
                m = _PAT_CODIGO_COLUMNA.search(columna)
                codigo = m.group(1).upper() if m else None
                valor = fila.resultados_por_codigo.get(codigo) if codigo and codigo in fila.analitos_solicitados else None
            ws.cell(row=fila_idx, column=col_idx, value=valor)

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=resultados_cromatografia.xlsx"},
    )


def _nombre_informe(campos: dict[str, str]) -> str:
    n_solicitud = (campos.get("N° Solicitud") or "solicitud").replace("/", "-")
    return f"Informe_{n_solicitud}.pdf"


def _archivar_informe(campos: dict[str, str], folio: str, pdf_bytes: bytes) -> None:
    """Guarda una copia del informe en R2, agrupado igual que las solicitudes
    -por cliente y día- pero bajo `informes/`, su propia raíz:

        informes/<SOLD TO>/<AAAA-MM-DD>/<folio>.pdf

    El archivado no puede costarle la descarga al usuario: si R2 no está
    configurado o falla, se registra y el informe se entrega igual.
    """
    if not r2.disponible():
        return
    cliente = carpeta_de_cliente(campos.get("Sold To (Nombre)"))
    fecha = date.today().isoformat()
    key = f"informes/{cliente}/{fecha}/{folio}.pdf"
    try:
        r2.subir(key, pdf_bytes, "application/pdf")
    except Exception:
        logging.getLogger(__name__).exception("No se pudo archivar el informe %s en R2", folio)


@router.post("/informes-pdf")
def generar_informes_pdf(filas: list[FilaCruceIn]) -> StreamingResponse:
    if not filas:
        raise HTTPException(400, "No hay filas para exportar.")

    with conexion() as conn, cursor_dict(conn) as cur:
        folios = _asignar_folios(cur, len(filas))
    config_fila = leer_config_informe()

    def _generar(fila: FilaCruceIn, folio: str) -> bytes:
        pdf_bytes = generar_informe_pdf(
            campos=fila.campos,
            analitos_solicitados=fila.analitos_solicitados,
            resultados_por_codigo=fila.resultados_por_codigo,
            codigo_vial=fila.codigo_vial,
            fecha_inyeccion=fila.fecha_inyeccion,
            fecha_recepcion=fila.fecha_recepcion,
            folio=folio,
            analizado_por_nombre=config_fila.get("analizado_por_nombre") or "",
            analizado_por_cargo=config_fila.get("analizado_por_cargo") or "",
            aprobado_por_nombre=config_fila.get("aprobado_por_nombre") or "",
            aprobado_por_cargo=config_fila.get("aprobado_por_cargo") or "",
            incluir_analista=config_fila.get("incluir_analista", True),
        )
        _archivar_informe(fila.campos, folio, pdf_bytes)
        return pdf_bytes

    if len(filas) == 1:
        fila = filas[0]
        pdf_bytes = _generar(fila, folios[0])
        return StreamingResponse(
            io.BytesIO(pdf_bytes),
            media_type="application/pdf",
            headers={"Content-Disposition": f'attachment; filename="{_nombre_informe(fila.campos)}"'},
        )

    # Varias solicitudes cruzadas a la vez: un PDF por cada una, empaquetados
    # en un único zip para descargar de una sola vez.
    buffer = io.BytesIO()
    usados: set[str] = set()
    with zipfile.ZipFile(buffer, "w", zipfile.ZIP_DEFLATED) as zf:
        for fila, folio in zip(filas, folios):
            pdf_bytes = _generar(fila, folio)
            base, ext = os.path.splitext(_nombre_informe(fila.campos))
            nombre, n = f"{base}{ext}", 2
            while nombre in usados:
                nombre, n = f"{base} ({n}){ext}", n + 1
            usados.add(nombre)
            zf.writestr(nombre, pdf_bytes)
    buffer.seek(0)
    return StreamingResponse(
        buffer,
        media_type="application/zip",
        headers={"Content-Disposition": 'attachment; filename="informes_cromatografia.zip"'},
    )


# ---------------------------------------------------------------------------
# Subir a base de datos: la fila cruzada (solicitud + resultado del GC ya
# validado, sin cruces sospechosos) pasa a ser un registro real de `solicitud`
# + `resultado`, visible en Report/DataCore igual que cualquier otro dato
# cargado por Ingest. Se identifica con NUESTRO folio (LAB-YYYYMMDD-NNN) en
# vez del N° de solicitud original -así el registro queda trazado al informe
# que se emitió, no al papeleo de origen-.
# ---------------------------------------------------------------------------

TIPO_SERVICIO_CROMATOGRAFIA = "Cromatografía"


def _parse_fecha(texto: str | None, formato: str) -> date | None:
    if not texto:
        return None
    try:
        return datetime.strptime(texto.strip(), formato).date()
    except ValueError:
        return None


def _fecha_ddmmyyyy(texto: str | None) -> date | None:
    """'18-08-2026' o '19-08-2026 14:59' -> solo la fecha, se descarta la hora
    (la tabla solicitud guarda DATE, no timestamp, para estos campos)."""
    if not texto:
        return None
    return _parse_fecha(texto, "%d-%m-%Y %H:%M") or _parse_fecha(texto, "%d-%m-%Y")


def _fecha_inyeccion_gc(texto: str | None) -> date | None:
    """Formato de Agilent ChemStation: '7/25/2026 9:14:59 AM'."""
    return _parse_fecha(texto, "%m/%d/%Y %I:%M:%S %p")


class FilaSubidaOut(BaseModel):
    nro_solicitud_original: str
    codigo_vial: str | None
    estado: str  # 'creada' | 'ya_existia' | 'error'
    folio: str | None = None
    mensaje: str | None = None


def _resolver_cliente_planta(cur, sold_to: str | None, ship_to: str | None) -> tuple[int | None, int | None, str | None]:
    """Nunca crea cliente/planta nuevos acá -eso es responsabilidad exclusiva
    del catálogo oficial (Listados, ver catalogo.py)-: si el Sold To/Ship To
    de la solicitud no calza exacto con algo ya cargado ahí, se rechaza con un
    mensaje claro en vez de inventar un cliente nuevo silenciosamente."""
    if not sold_to:
        return None, None, "La solicitud no trae Sold To: no se puede subir a la base de datos."
    cur.execute("SELECT id FROM cliente WHERE nombre = %s AND activo", (sold_to,))
    cliente = cur.fetchone()
    if not cliente:
        return None, None, f'Sold To "{sold_to}" no está en el catálogo (Listados). Agrégalo ahí primero.'
    cliente_id = cliente["id"]
    if not ship_to:
        return cliente_id, None, None
    cur.execute("SELECT id FROM planta WHERE cliente_id = %s AND nombre = %s AND activo", (cliente_id, ship_to))
    planta = cur.fetchone()
    if not planta:
        return None, None, f'Ship To "{ship_to}" no está en el catálogo (Listados) para el Sold To "{sold_to}". Agrégalo ahí primero.'
    return cliente_id, planta["id"], None


@router.post("/subir-bd")
def subir_bd(filas: list[FilaCruceIn]) -> list[FilaSubidaOut]:
    if not filas:
        raise HTTPException(400, "No hay filas para subir.")

    salida: list[FilaSubidaOut] = []
    con_folio: list[tuple[FilaCruceIn, str, int | None]] = []

    with conexion(escribir=True) as conn, cursor_dict(conn) as cur:
        for fila in filas:
            nro_original = fila.campos.get("N° Solicitud") or "—"
            codigo_vial = fila.codigo_vial

            # Ya se subió antes: misma solicitud original + mismo vial ya tiene
            # un registro (referencia/nro_orden), así que no se duplica -clic
            # repetido al botón, o volver a cruzar lo mismo por accidente-.
            cur.execute(
                "SELECT nro_solicitud FROM solicitud WHERE referencia = %s AND nro_orden = %s AND laboratorio = %s",
                (nro_original, codigo_vial, LABORATORIO_CATALOGO),
            )
            existente = cur.fetchone()
            if existente:
                salida.append(
                    FilaSubidaOut(
                        nro_solicitud_original=nro_original,
                        codigo_vial=codigo_vial,
                        estado="ya_existia",
                        folio=existente["nro_solicitud"],
                        mensaje="Esta solicitud y vial ya se habían subido antes; no se duplicó.",
                    )
                )
                continue

            cliente_id, planta_id, error = _resolver_cliente_planta(
                cur, fila.campos.get("Sold To (Nombre)"), fila.campos.get("Ship To (Nombre)")
            )
            if error:
                salida.append(
                    FilaSubidaOut(nro_solicitud_original=nro_original, codigo_vial=codigo_vial, estado="error", mensaje=error)
                )
                continue

            con_folio.append((fila, nro_original, planta_id))

        if con_folio:
            folios = _asignar_folios(cur, len(con_folio))
            for (fila, nro_original, planta_id), folio in zip(con_folio, folios):
                fecha_muestreo = _fecha_ddmmyyyy(fila.campos.get("Fecha Muestreo"))
                datos_solicitud = {
                    "nro_solicitud": folio,
                    "laboratorio": LABORATORIO_CATALOGO,
                    "fecha_solicitud": _fecha_ddmmyyyy(fila.campos.get("Fecha Solicitud")),
                    "fecha_muestreo": fecha_muestreo,
                    "fecha_analisis": _fecha_inyeccion_gc(fila.fecha_inyeccion),
                    "fecha_recepcion": _parse_fecha(fila.fecha_recepcion, "%Y-%m-%d"),
                    "fecha_informe": date.today(),
                    "planta_id": planta_id,
                    "sold_to_raw": fila.campos.get("Sold To (Nombre)"),
                    "ship_to_raw": fila.campos.get("Ship To (Nombre)") or None,
                    "especie": fila.campos.get("Especie") or None,
                    "variedad": fila.campos.get("Variedad") or None,
                    "tipo_muestra": fila.campos.get("Tipo Muestra") or None,
                    "tipo_servicio": TIPO_SERVICIO_CROMATOGRAFIA,
                    "lote": fila.campos.get("Lote") or None,
                    "solicitante": fila.campos.get("Solicitante") or None,
                    "nombre_muestreador": fila.campos.get("Nombre Muestreador") or None,
                    "generado_por": fila.campos.get("Generado Por") or None,
                    "email_solicitante": fila.campos.get("Email Solicitante") or None,
                    "nro_orden": fila.codigo_vial,
                    "referencia": nro_original,
                    "observacion": fila.campos.get("Producto Utilizado") or None,
                    "semana_muestreo": calcular_semana(fecha_muestreo.isoformat()) if fecha_muestreo else None,
                    "mes": fecha_muestreo.month if fecha_muestreo else None,
                    "origen": "emitir_cromatografia",
                }
                columnas = list(datos_solicitud.keys())
                placeholders = ", ".join(["%s"] * len(columnas))
                cur.execute(
                    f"INSERT INTO solicitud ({', '.join(columnas)}) VALUES ({placeholders}) RETURNING id",
                    [datos_solicitud[c] for c in columnas],
                )
                solicitud_id = cur.fetchone()["id"]

                for codigo in fila.analitos_solicitados:
                    valor = fila.resultados_por_codigo.get(codigo)
                    cur.execute("SELECT id FROM analito WHERE codigo = %s AND laboratorio = %s", (codigo, LABORATORIO_CATALOGO))
                    analito = cur.fetchone()
                    cur.execute(
                        """INSERT INTO resultado (solicitud_id, analito_id, analito_raw, valor_num)
                           VALUES (%s, %s, %s, %s)
                           ON CONFLICT (solicitud_id, analito_id) DO NOTHING""",
                        (solicitud_id, analito["id"] if analito else None, None if analito else codigo, valor),
                    )

                salida.append(
                    FilaSubidaOut(nro_solicitud_original=nro_original, codigo_vial=fila.codigo_vial, estado="creada", folio=folio)
                )

    return salida
