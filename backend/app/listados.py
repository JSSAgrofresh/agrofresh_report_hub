"""
Mantenedor "Listados": fuente de verdad de los valores estandarizados y
seleccionables de Especie y Variedad (Sold To / Ship To siguen viviendo en
cliente/planta -ver catalogo.py-, ya son la fuente de verdad de esos dos y no
se duplican acá).

Modelo de datos (valor_lista, tipo especie/variedad):
- Valor "crudo": una fila normal (es_estandar=false). Si `fusionado_en_id` es
  NULL, es un valor seleccionable tal cual. Si apunta a otra fila, significa
  que un administrador lo asignó a esa variedad estandarizada -queda inactivo
  para no aparecer duplicado en los selects, pero NUNCA se borra: las
  solicitudes históricas guardan el texto tal cual, no un ID-.
- Variedad "estándar": una fila con es_estandar=true. Es el valor que
  realmente ofrecen los selects de la app junto con los valores crudos sin
  asignar. Se crea, renombra y elimina a mano desde /estandares.
- Toda fila tipo=variedad tiene `especie_id` (apunta a una fila tipo=especie):
  el mismo texto puede existir en más de una especie (ej. "June Gold" es una
  variedad de Durazno Y, por separado, de Manzana) y NO deben fusionarse
  entre sí -por eso Variedad ya no es una lista plana, cada valor pertenece a
  una especie concreta-.

"Homogenizar" (GET /{tipo}/homogenizar) NUNCA fusiona nada: solo agrupa
valores crudos que probablemente son el mismo dato mal escrito (mayúsculas,
acentos, espacios, puntuación -alta confianza-, o variantes ortográficas
obvias -a revisar-) como ayuda de revisión, siempre DENTRO de la misma
especie para variedad. Un mismo grupo de similitud puede contener MÁS DE UNA
variedad real (ej. "Packham" y "Packham's Triumph" caen en el mismo grupo por
nombre parecido, pero son variedades distintas), así que el administrador
decide libremente cuántas variedades estándar crea a partir de un grupo y qué
valores le asigna a cada una -ver /estandares y /{tipo}/{id}/asignar-.
"""
import io
import re
import unicodedata
from collections import Counter
from datetime import datetime
from difflib import SequenceMatcher
from typing import Any

from fastapi import APIRouter, HTTPException, Query
from fastapi.responses import StreamingResponse
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.worksheet.worksheet import Worksheet
from pydantic import BaseModel

from .db import conexion, cursor_dict

router = APIRouter(prefix="/api/listados", tags=["listados"])

TIPOS_VALIDOS = ("especie", "variedad")


def _validar_tipo(tipo: str) -> str:
    if tipo not in TIPOS_VALIDOS:
        raise HTTPException(404, f"Listado '{tipo}' no existe. Tipos disponibles: {', '.join(TIPOS_VALIDOS)}")
    return tipo


def normalizar_texto_general(valor: str) -> str:
    """"uva"/"UVA"/"uVa" -> "Uva". Solo para listas de texto general
    (Especie, Variedad) -NUNCA para Sold To/Ship To, que se guardan tal cual
    los entrega el cliente o SAP-."""
    limpio = re.sub(r"\s+", " ", (valor or "").strip())
    return limpio.title() if limpio else limpio


def clave_normalizada(valor: str) -> str:
    """Clave de agrupación insensible a mayúsculas, acentos, espacios y
    puntuación: 'Thompson', 'THOMPSON' y ' thompson  ' comparten la misma
    clave, pero 'Thompson' y 'Thompson Seedless' NO (esas se detectan aparte,
    como candidatas "a revisar", no como duplicado automático)."""
    v = unicodedata.normalize("NFKD", valor or "")
    v = "".join(c for c in v if not unicodedata.combining(c))
    v = re.sub(r"[^a-z0-9]+", " ", v.lower()).strip()
    return re.sub(r"\s+", " ", v)


class ValorListaIn(BaseModel):
    valor: str
    activo: bool = True
    # Obligatorio cuando tipo=variedad; se ignora para tipo=especie.
    especie_id: int | None = None


class AsignarIn(BaseModel):
    estandar_id: int | None = None


def _validar_especie_id(cur, tipo: str, especie_id: int | None) -> int | None:
    """Variedad SIEMPRE necesita una especie; Especie nunca lleva una."""
    if tipo == "especie":
        return None
    if especie_id is None:
        raise HTTPException(400, "Elige a qué especie pertenece esta variedad.")
    cur.execute("SELECT 1 FROM lab.valor_lista WHERE id = %s AND tipo = 'especie'", (especie_id,))
    if not cur.fetchone():
        raise HTTPException(404, "La especie indicada no existe.")
    return especie_id


def _buscar_o_crear_estandar(cur, tipo: str, valor_crudo: str, especie_id: int | None) -> int:
    """Encuentra la variedad estándar con ese nombre EN ESA ESPECIE (para que
    crear "Packham" desde dos grupos de similitud distintos de la misma
    especie termine en la MISMA fila, pero un "August" de Durazno nunca se
    confunda con un "August" de Nectarina) o la crea si no existe. Caso
    normal: el nombre elegido para la variedad estándar coincide con uno de
    los valores crudos que se le están por asignar -esa fila se "promueve" a
    variedad estándar en vez de tratarse como un choque-."""
    valor = normalizar_texto_general(valor_crudo)
    clave = clave_normalizada(valor)
    if tipo == "especie":
        cur.execute("SELECT id, es_estandar FROM lab.valor_lista WHERE tipo = 'especie' AND valor_normalizado = %s", (clave,))
    else:
        cur.execute(
            "SELECT id, es_estandar FROM lab.valor_lista WHERE tipo = 'variedad' AND especie_id = %s AND valor_normalizado = %s",
            (especie_id, clave),
        )
    existente = cur.fetchone()
    if existente:
        if not existente["es_estandar"]:
            cur.execute(
                "UPDATE lab.valor_lista SET es_estandar = true, activo = true, fusionado_en_id = NULL, valor = %s WHERE id = %s",
                (valor, existente["id"]),
            )
        return existente["id"]
    cur.execute(
        "INSERT INTO lab.valor_lista (tipo, valor, valor_normalizado, activo, es_estandar, especie_id) VALUES (%s, %s, %s, true, true, %s) RETURNING id",
        (tipo, valor, clave, especie_id),
    )
    return cur.fetchone()["id"]


_VERDE_OSCURO = "3D6B1F"
_FUENTE_HEADER = Font(bold=True, color="FFFFFF", size=10.5)
_RELLENO_HEADER = PatternFill("solid", fgColor=_VERDE_OSCURO)


def _escribir_hoja(wb: Workbook, titulo: str, encabezados: list[str], filas: list[list[Any]], anchos: list[int]) -> Worksheet:
    ws = wb.create_sheet(titulo[:31])
    for col, texto in enumerate(encabezados, start=1):
        c = ws.cell(row=1, column=col, value=texto)
        c.font = _FUENTE_HEADER
        c.fill = _RELLENO_HEADER
        c.alignment = Alignment(vertical="center")
    ws.row_dimensions[1].height = 20
    for fila_idx, fila in enumerate(filas, start=2):
        for col_idx, valor in enumerate(fila, start=1):
            ws.cell(row=fila_idx, column=col_idx, value=valor)
    for col_idx, ancho in enumerate(anchos, start=1):
        ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = ancho
    ws.freeze_panes = "A2"
    if filas:
        ws.auto_filter.ref = f"A1:{ws.cell(row=1, column=len(encabezados)).column_letter}{len(filas) + 1}"
    return ws


@router.get("/exportar")
def exportar_listados() -> StreamingResponse:
    """Un único Excel con las 4 listas -Sold To, Ship To, Especie, Variedad-,
    tal como quedan después de homogenizar: variedades estándar primero, y
    para cada valor crudo a qué variedad estándar quedó asignado (o vacío si
    todavía está pendiente). Variedad incluye la Especie a la que pertenece."""
    with conexion() as conn, cursor_dict(conn) as cur:
        cur.execute("SELECT nombre, codigo_sap, rut, activo FROM cliente ORDER BY nombre")
        clientes = cur.fetchall()
        cur.execute(
            "SELECT p.nombre, p.codigo_sap, c.nombre AS cliente_nombre, p.ciudad, p.activo "
            "FROM planta p JOIN cliente c ON c.id = p.cliente_id ORDER BY c.nombre, p.nombre"
        )
        plantas = cur.fetchall()

        cur.execute(
            "SELECT valor, activo, es_estandar, NULL AS asignado_a "
            "FROM lab.valor_lista WHERE tipo = 'especie' ORDER BY es_estandar DESC, valor"
        )
        especies = cur.fetchall()
        cur.execute(
            "SELECT v.valor, v.activo, v.es_estandar, esp.valor AS especie, e.valor AS asignado_a "
            "FROM lab.valor_lista v "
            "JOIN lab.valor_lista esp ON esp.id = v.especie_id "
            "LEFT JOIN lab.valor_lista e ON e.id = v.fusionado_en_id "
            "WHERE v.tipo = 'variedad' ORDER BY esp.valor, v.es_estandar DESC, v.valor"
        )
        variedades = cur.fetchall()

    wb = Workbook()
    wb.remove(wb.active)

    _escribir_hoja(
        wb,
        "Sold To",
        ["N° Sold To", "Sold To", "RUT", "Estado"],
        [[c["codigo_sap"], c["nombre"], c["rut"], "Activo" if c["activo"] else "Inactivo"] for c in clientes],
        [14, 44, 16, 12],
    )
    _escribir_hoja(
        wb,
        "Ship To",
        ["N° Ship To", "Ship To", "Sold To", "Ciudad", "Estado"],
        [[p["codigo_sap"], p["nombre"], p["cliente_nombre"], p["ciudad"], "Activo" if p["activo"] else "Inactivo"] for p in plantas],
        [14, 44, 34, 20, 12],
    )
    _escribir_hoja(
        wb,
        "Especie",
        ["Especie", "Tipo", "Estado", "Estándar asignado"],
        [[f["valor"], "Estándar" if f["es_estandar"] else "Crudo", "Activo" if f["activo"] else "Inactivo", f["asignado_a"] or ""] for f in especies],
        [30, 12, 12, 30],
    )
    _escribir_hoja(
        wb,
        "Variedad",
        ["Especie", "Variedad", "Tipo", "Estado", "Variedad estándar asignada"],
        [
            [f["especie"], f["valor"], "Estándar" if f["es_estandar"] else "Crudo", "Activo" if f["activo"] else "Inactivo", f["asignado_a"] or ""]
            for f in variedades
        ],
        [18, 30, 12, 12, 30],
    )

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    nombre_archivo = f"listados_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{nombre_archivo}"'},
    )


@router.get("/{tipo}")
def listar_valores(
    tipo: str,
    incluir_inactivos: bool = Query(False),
    buscar: str | None = Query(None),
    especie_id: int | None = Query(None),
) -> list[dict[str, Any]]:
    _validar_tipo(tipo)
    with conexion() as conn, cursor_dict(conn) as cur:
        sql = "SELECT id, tipo, valor, activo, es_estandar, fusionado_en_id, especie_id, creado_en FROM lab.valor_lista WHERE tipo = %s"
        params: list[Any] = [tipo]
        if not incluir_inactivos:
            sql += " AND activo = true"
        if buscar:
            sql += " AND valor ILIKE %s"
            params.append(f"%{buscar}%")
        if tipo == "variedad" and especie_id is not None:
            sql += " AND especie_id = %s"
            params.append(especie_id)
        sql += " ORDER BY valor"
        cur.execute(sql, params)
        return cur.fetchall()


@router.post("/{tipo}")
def crear_valor(tipo: str, body: ValorListaIn) -> dict[str, Any]:
    _validar_tipo(tipo)
    valor = normalizar_texto_general(body.valor)
    if not valor:
        raise HTTPException(400, "El valor no puede estar vacío.")
    clave = clave_normalizada(valor)
    with conexion() as conn, cursor_dict(conn) as cur:
        especie_id = _validar_especie_id(cur, tipo, body.especie_id)
        if tipo == "especie":
            cur.execute("SELECT id FROM lab.valor_lista WHERE tipo = 'especie' AND valor_normalizado = %s", (clave,))
        else:
            cur.execute(
                "SELECT id FROM lab.valor_lista WHERE tipo = 'variedad' AND especie_id = %s AND valor_normalizado = %s",
                (especie_id, clave),
            )
        if cur.fetchone():
            raise HTTPException(409, f"Ya existe un valor equivalente en {tipo}.")
        cur.execute(
            "INSERT INTO lab.valor_lista (tipo, valor, valor_normalizado, activo, especie_id) VALUES (%s, %s, %s, %s, %s) RETURNING id",
            (tipo, valor, clave, body.activo, especie_id),
        )
        return {"id": cur.fetchone()["id"]}


@router.put("/{tipo}/{valor_id}")
def editar_valor(tipo: str, valor_id: int, body: ValorListaIn) -> dict[str, str]:
    _validar_tipo(tipo)
    valor = normalizar_texto_general(body.valor)
    if not valor:
        raise HTTPException(400, "El valor no puede estar vacío.")
    clave = clave_normalizada(valor)
    with conexion() as conn, cursor_dict(conn) as cur:
        if tipo == "especie":
            especie_id = None
            cur.execute(
                "SELECT id FROM lab.valor_lista WHERE tipo = 'especie' AND valor_normalizado = %s AND id != %s",
                (clave, valor_id),
            )
        else:
            cur.execute("SELECT id FROM lab.valor_lista WHERE id = %s AND tipo = 'variedad'", (valor_id,))
            if not cur.fetchone():
                raise HTTPException(404, "Valor no encontrado")
            # A diferencia de crear_valor, acá SÍ se permite reasignar la
            # especie -es la vía para corregir a mano una variedad que quedó
            # sin especie (datos legados) o mal clasificada-.
            especie_id = _validar_especie_id(cur, tipo, body.especie_id)
            cur.execute(
                "SELECT id FROM lab.valor_lista WHERE tipo = 'variedad' AND especie_id = %s AND valor_normalizado = %s AND id != %s",
                (especie_id, clave, valor_id),
            )
        if cur.fetchone():
            raise HTTPException(409, f"Ya existe otro valor equivalente en {tipo}.")
        cur.execute(
            "UPDATE lab.valor_lista SET valor = %s, valor_normalizado = %s, activo = %s, especie_id = %s WHERE id = %s AND tipo = %s",
            (valor, clave, body.activo, especie_id, valor_id, tipo),
        )
        if cur.rowcount == 0:
            raise HTTPException(404, "Valor no encontrado")
        return {"estado": "ok"}


@router.delete("/{tipo}/{valor_id}")
def eliminar_valor(tipo: str, valor_id: int) -> dict[str, str]:
    """Borrado FÍSICO y definitivo -a propósito, se reserva para variedades
    estándar vacías (creadas de más, sin nada asignado). Un valor crudo
    (es_estandar=false) NUNCA se borra físicamente acá: si sobra, se
    desactiva (PUT activo=false); si estaba asignado y se desasigna, tiene
    que volver disponible para homogenizar de nuevo -no desaparecer-. Borrar
    en la tabla general del mantenedor rompía justamente eso."""
    _validar_tipo(tipo)
    with conexion() as conn, cursor_dict(conn) as cur:
        cur.execute("SELECT es_estandar FROM lab.valor_lista WHERE id = %s AND tipo = %s", (valor_id, tipo))
        fila = cur.fetchone()
        if not fila:
            raise HTTPException(404, "Valor no encontrado")
        if not fila["es_estandar"]:
            raise HTTPException(
                409,
                "Un valor crudo no se elimina físicamente: desactívalo, o si está asignado, desasígnalo desde "
                "Homogenizar para que vuelva a la lista pendiente.",
            )
        if tipo == "especie":
            cur.execute("SELECT 1 FROM lab.valor_lista WHERE especie_id = %s", (valor_id,))
            if cur.fetchone():
                raise HTTPException(409, "Esta especie todavía tiene variedades asociadas: no se puede eliminar.")
        cur.execute("SELECT 1 FROM lab.valor_lista WHERE fusionado_en_id = %s", (valor_id,))
        if cur.fetchone():
            raise HTTPException(
                409,
                "Esta variedad estándar todavía tiene valores asignados: elimínala desde Homogenizar para "
                "liberarlos primero.",
            )
        cur.execute("DELETE FROM lab.valor_lista WHERE id = %s AND tipo = %s", (valor_id, tipo))
        if cur.rowcount == 0:
            raise HTTPException(404, "Valor no encontrado")
        return {"estado": "ok"}


@router.get("/{tipo}/homogenizar")
def candidatos_homogenizacion(tipo: str, especie_id: int | None = Query(None)) -> list[dict[str, Any]]:
    """Agrupa valores activos que probablemente son el mismo dato repetido.
    Nunca fusiona nada solo: solo propone -ver /estandares y /asignar-. Para
    variedad, especie_id es obligatorio: nunca se agrupan valores de
    especies distintas, aunque el texto sea idéntico (ej. "June Gold" de
    Durazno y "June Gold" de Manzana son variedades distintas)."""
    _validar_tipo(tipo)
    if tipo == "variedad" and especie_id is None:
        raise HTTPException(400, "Elige una especie primero para homogenizar sus variedades.")
    with conexion() as conn, cursor_dict(conn) as cur:
        sql = "SELECT id, valor FROM lab.valor_lista WHERE tipo = %s AND activo = true AND es_estandar = false AND fusionado_en_id IS NULL"
        params: list[Any] = [tipo]
        if tipo == "variedad":
            sql += " AND especie_id = %s"
            params.append(especie_id)
        sql += " ORDER BY valor"
        cur.execute(sql, params)
        filas = cur.fetchall()

    for f in filas:
        f["_clave"] = clave_normalizada(f["valor"])

    # Etapa 1 (alta confianza): misma clave normalizada -difieren solo en
    # mayúsculas, acentos, espacios o puntuación-.
    por_clave: dict[str, list[dict]] = {}
    for f in filas:
        por_clave.setdefault(f["_clave"], []).append(f)

    grupos: list[dict[str, Any]] = []
    usados_ids: set[int] = set()
    for miembros in por_clave.values():
        if len(miembros) > 1:
            propuesto = Counter(m["valor"] for m in miembros).most_common(1)[0][0]
            grupos.append(
                {
                    "confianza": "alta",
                    "valores": [{"id": m["id"], "valor": m["valor"]} for m in miembros],
                    "valor_propuesto": normalizar_texto_general(propuesto),
                }
            )
            usados_ids.update(m["id"] for m in miembros)

    # Etapa 2 (a revisar): variantes ortográficas obvias entre claves
    # DISTINTAS, nunca solo "se parecen un poco". Dos criterios, cada uno
    # pensado para no confundir nombres realmente distintos:
    #   (a) una clave es, palabra por palabra, PREFIJO completo de la otra
    #       -"Thompson" de "Thompson Seedless"-: nunca por substring suelto,
    #       porque eso encadenaría cualquier cosa que comparta un prefijo
    #       corto y genérico (ej. "Pears", "Summer").
    #   (b) mismo número de palabras y altísima similitud de caracteres
    #       -typos como "Honey Crips"/"Honey Crisp"-. El umbral es alto a
    #       propósito: a 0.90 ya agrupaba "Gala"/"Galaxy", que son
    #       variedades distintas, no un typo.
    # Se agrupan por conjuntos disjuntos (union-find) para juntar cadenas de
    # variantes (A~B~C).
    restantes = [f for f in filas if f["id"] not in usados_ids]
    padre = {f["id"]: f["id"] for f in restantes}

    def encontrar(x: int) -> int:
        while padre[x] != x:
            padre[x] = padre[padre[x]]
            x = padre[x]
        return x

    def unir(a: int, b: int) -> None:
        ra, rb = encontrar(a), encontrar(b)
        if ra != rb:
            padre[ra] = rb

    LARGO_MIN_PREFIJO = 6
    UMBRAL_TIPO = 0.93
    for i in range(len(restantes)):
        for j in range(i + 1, len(restantes)):
            a, b = restantes[i], restantes[j]
            if not a["_clave"] or not b["_clave"]:
                continue
            ta, tb = a["_clave"].split(" "), b["_clave"].split(" ")
            corta, larga = (ta, tb) if len(ta) <= len(tb) else (tb, ta)
            es_prefijo = (
                len(corta) < len(larga)
                and larga[: len(corta)] == corta
                and len(" ".join(corta)) >= LARGO_MIN_PREFIJO
            )
            ratio = SequenceMatcher(None, a["_clave"], b["_clave"]).ratio()
            mismo_n_palabras = len(ta) == len(tb)
            if es_prefijo or (mismo_n_palabras and ratio >= UMBRAL_TIPO):
                unir(a["id"], b["id"])

    por_raiz: dict[int, list[dict]] = {}
    for f in restantes:
        por_raiz.setdefault(encontrar(f["id"]), []).append(f)

    for miembros in por_raiz.values():
        if len(miembros) > 1:
            propuesto = min(miembros, key=lambda m: len(m["valor"]))["valor"]
            grupos.append(
                {
                    "confianza": "revisar",
                    "valores": [{"id": m["id"], "valor": m["valor"]} for m in miembros],
                    "valor_propuesto": normalizar_texto_general(propuesto),
                }
            )

    grupos.sort(key=lambda g: (g["confianza"] != "alta", -len(g["valores"])))
    return grupos


@router.get("/{tipo}/estandares")
def listar_estandares(tipo: str, especie_id: int | None = Query(None)) -> dict[str, Any]:
    """Cada variedad estándar con los valores crudos que un administrador le
    asignó, más los valores crudos activos que todavía no se asignaron a
    ninguna. Es la vista de "clasificación final" -a diferencia de
    /homogenizar, que es solo la ayuda de revisión-. Para variedad,
    especie_id es obligatorio -mismo motivo que en /homogenizar-."""
    _validar_tipo(tipo)
    if tipo == "variedad" and especie_id is None:
        raise HTTPException(400, "Elige una especie primero.")
    filtro_especie = " AND especie_id = %s" if tipo == "variedad" else ""
    params_especie: list[Any] = [especie_id] if tipo == "variedad" else []
    with conexion() as conn, cursor_dict(conn) as cur:
        cur.execute(
            f"SELECT id, valor, activo FROM lab.valor_lista WHERE tipo = %s AND es_estandar = true{filtro_especie} ORDER BY valor",
            [tipo, *params_especie],
        )
        estandares = cur.fetchall()
        cur.execute(
            f"SELECT id, valor, fusionado_en_id FROM lab.valor_lista WHERE tipo = %s AND es_estandar = false AND fusionado_en_id IS NOT NULL{filtro_especie} ORDER BY valor",
            [tipo, *params_especie],
        )
        asignados = cur.fetchall()
        cur.execute(
            f"SELECT id, valor FROM lab.valor_lista WHERE tipo = %s AND es_estandar = false AND fusionado_en_id IS NULL AND activo = true{filtro_especie} ORDER BY valor",
            [tipo, *params_especie],
        )
        sin_asignar = cur.fetchall()

    por_estandar: dict[int, list[dict]] = {}
    for a in asignados:
        por_estandar.setdefault(a["fusionado_en_id"], []).append({"id": a["id"], "valor": a["valor"]})

    return {
        "estandares": [
            {"id": e["id"], "valor": e["valor"], "activo": e["activo"], "valores_asignados": por_estandar.get(e["id"], [])}
            for e in estandares
        ],
        "sin_asignar": [{"id": s["id"], "valor": s["valor"]} for s in sin_asignar],
    }


@router.post("/{tipo}/estandares")
def crear_estandar(tipo: str, body: ValorListaIn) -> dict[str, Any]:
    """Crea una variedad estándar con nombre completamente libre -no tiene
    que derivarse del valor más común de ningún grupo-. Si el administrador
    reutiliza un nombre que ya existe como estándar EN LA MISMA ESPECIE, se
    reusa esa misma fila en vez de duplicarla."""
    _validar_tipo(tipo)
    if not body.valor.strip():
        raise HTTPException(400, "El nombre de la variedad estándar no puede estar vacío.")
    with conexion() as conn, cursor_dict(conn) as cur:
        especie_id = _validar_especie_id(cur, tipo, body.especie_id)
        estandar_id = _buscar_o_crear_estandar(cur, tipo, body.valor, especie_id)
        return {"id": estandar_id}


@router.put("/{tipo}/estandares/{estandar_id}")
def editar_estandar(tipo: str, estandar_id: int, body: ValorListaIn) -> dict[str, str]:
    _validar_tipo(tipo)
    valor = normalizar_texto_general(body.valor)
    if not valor:
        raise HTTPException(400, "El valor no puede estar vacío.")
    clave = clave_normalizada(valor)
    with conexion() as conn, cursor_dict(conn) as cur:
        cur.execute("SELECT especie_id FROM lab.valor_lista WHERE id = %s AND tipo = %s AND es_estandar = true", (estandar_id, tipo))
        actual = cur.fetchone()
        if not actual:
            raise HTTPException(404, "Variedad estándar no encontrada")
        if tipo == "especie":
            cur.execute(
                "SELECT id FROM lab.valor_lista WHERE tipo = 'especie' AND valor_normalizado = %s AND id != %s",
                (clave, estandar_id),
            )
        else:
            cur.execute(
                "SELECT id FROM lab.valor_lista WHERE tipo = 'variedad' AND especie_id = %s AND valor_normalizado = %s AND id != %s",
                (actual["especie_id"], clave, estandar_id),
            )
        if cur.fetchone():
            raise HTTPException(409, f"Ya existe otro valor equivalente en {tipo}.")
        cur.execute(
            "UPDATE lab.valor_lista SET valor = %s, valor_normalizado = %s, activo = %s WHERE id = %s AND tipo = %s AND es_estandar = true",
            (valor, clave, body.activo, estandar_id, tipo),
        )
        if cur.rowcount == 0:
            raise HTTPException(404, "Variedad estándar no encontrada")
        return {"estado": "ok"}


@router.delete("/{tipo}/estandares/{estandar_id}")
def eliminar_estandar(tipo: str, estandar_id: int) -> dict[str, str]:
    """Elimina la variedad estándar y libera a todos los valores crudos que
    tenía asignados -vuelven a quedar activos y sin asignar, no se borran-."""
    _validar_tipo(tipo)
    with conexion() as conn, cursor_dict(conn) as cur:
        cur.execute(
            "UPDATE lab.valor_lista SET activo = true, fusionado_en_id = NULL WHERE tipo = %s AND fusionado_en_id = %s",
            (tipo, estandar_id),
        )
        cur.execute("DELETE FROM lab.valor_lista WHERE id = %s AND tipo = %s AND es_estandar = true", (estandar_id, tipo))
        if cur.rowcount == 0:
            raise HTTPException(404, "Variedad estándar no encontrada")
        return {"estado": "ok"}


@router.post("/{tipo}/{valor_id}/asignar")
def asignar_valor(tipo: str, valor_id: int, body: AsignarIn) -> dict[str, str]:
    """Asigna (o desasigna, con estandar_id=null) un valor crudo a una
    variedad estándar. Es la operación atómica detrás de todo el flujo:
    "crear variedad(es) libremente desde un grupo de similitud" es, para el
    backend, una variedad nueva + N llamadas a este endpoint. Para variedad,
    el destino tiene que ser de la MISMA especie que el valor crudo."""
    _validar_tipo(tipo)
    with conexion() as conn, cursor_dict(conn) as cur:
        cur.execute(
            "SELECT id, es_estandar, especie_id FROM lab.valor_lista WHERE id = %s AND tipo = %s",
            (valor_id, tipo),
        )
        fila = cur.fetchone()
        if not fila:
            raise HTTPException(404, "Valor no encontrado")
        if fila["es_estandar"]:
            # Puede pasar sin querer: el valor que se está "asignando" es el
            # mismo que acaba de promoverse a variedad estándar (ver
            # _buscar_o_crear_estandar). Asignarlo a sí mismo es un no-op.
            if body.estandar_id == valor_id:
                return {"estado": "ok"}
            raise HTTPException(400, "Una variedad estándar no se puede asignar a otra.")

        if body.estandar_id is None:
            cur.execute(
                "UPDATE lab.valor_lista SET activo = true, fusionado_en_id = NULL WHERE id = %s",
                (valor_id,),
            )
            return {"estado": "ok"}

        cur.execute(
            "SELECT especie_id FROM lab.valor_lista WHERE id = %s AND tipo = %s AND es_estandar = true",
            (body.estandar_id, tipo),
        )
        destino = cur.fetchone()
        if not destino:
            raise HTTPException(404, "La variedad estándar de destino no existe.")
        if tipo == "variedad" and destino["especie_id"] != fila["especie_id"]:
            raise HTTPException(400, "No se puede asignar una variedad a una variedad estándar de otra especie.")
        cur.execute(
            "UPDATE lab.valor_lista SET activo = false, fusionado_en_id = %s WHERE id = %s",
            (body.estandar_id, valor_id),
        )
        return {"estado": "ok"}
