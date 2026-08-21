"""
Importa/actualiza los Listados (Sold To, Ship To, Especie, Variedad) desde el
Excel oficial de dos hojas entregado por AgroFresh.

- Hoja "ShIP TO SOLD TO": upsert en cliente/planta (mismo catálogo que ya usa
  Ingest/Converter -ver catalogo.py-). Nunca se toca el texto de Sold To/Ship
  To -se guarda tal cual viene del Excel, solo se recortan espacios-. Si el
  cliente o la planta ya existían, solo se completan RUT/ciudad cuando
  faltaban; el nombre nunca se pisa.
- Hoja "ESPECIE VARIEDAD": columna Especies -> valor_lista tipo='especie';
  columna Variedad -> valor_lista tipo='variedad'. Ambas se normalizan a
  "Primera Letra Mayúscula" y se agrupan por clave normalizada -mismo valor
  repetido con mayúsculas/acentos/espacios/puntuación distintos se colapsa
  en una sola fila-. Variantes ortográficas MÁS obvias (ej. "Abate Fetel" vs
  "Abatete Fetel") NO se fusionan acá: quedan como filas separadas para que
  el administrador las revise con "Homogenizar" en el mantenedor.

Seguro de repetir (idempotente): vuelve a correr sin duplicar filas.

Uso:
    cd backend && python3 scripts/importar_listados_excel.py [ruta_excel.xlsx]

Si no se pasa ruta, usa scripts/data/listados_seed.xlsx (copia del Excel
entregado por AgroFresh, agosto 2026).
"""
import os
import sys
from collections import Counter

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

from openpyxl import load_workbook  # noqa: E402

from app.db import conexion, cursor_dict  # noqa: E402
from app.listados import clave_normalizada, normalizar_texto_general  # noqa: E402

RUTA_POR_DEFECTO = os.path.join(os.path.dirname(os.path.abspath(__file__)), "data", "listados_seed.xlsx")
HOJA_CLIENTES = "ShIP TO SOLD TO"
HOJA_ESPECIE_VARIEDAD = "ESPECIE VARIEDAD"


def _texto(valor) -> str:
    return str(valor).strip() if valor not in (None, "") else ""


def importar_sold_to_ship_to(cur, wb) -> tuple[int, int]:
    ws = wb[HOJA_CLIENTES]
    filas = list(ws.iter_rows(values_only=True))[1:]  # salta encabezado

    clientes_nuevos = 0
    plantas_nuevas = 0
    cliente_id_por_codigo: dict[str, int] = {}

    for fila in filas:
        if len(fila) < 6:
            continue
        n_sold_to, sold_to, rut, n_ship_to, ship_to, ciudad = (_texto(v) for v in fila[:6])
        if not sold_to:
            continue

        cliente_id = cliente_id_por_codigo.get(n_sold_to)
        if cliente_id is None:
            cur.execute(
                "SELECT id, rut FROM cliente WHERE codigo_sap = %s OR nombre = %s LIMIT 1",
                (n_sold_to or None, sold_to),
            )
            fila_cliente = cur.fetchone()
            if fila_cliente:
                cliente_id = fila_cliente["id"]
                if rut and not fila_cliente["rut"]:
                    cur.execute("UPDATE cliente SET rut = %s WHERE id = %s", (rut, cliente_id))
            else:
                cur.execute(
                    "INSERT INTO cliente (nombre, codigo_sap, rut, activo) VALUES (%s, %s, %s, true) RETURNING id",
                    (sold_to, n_sold_to or None, rut or None),
                )
                cliente_id = cur.fetchone()["id"]
                clientes_nuevos += 1
            cliente_id_por_codigo[n_sold_to] = cliente_id

        if not ship_to:
            continue
        cur.execute(
            "SELECT id, ciudad FROM planta WHERE (codigo_sap = %s AND codigo_sap IS NOT NULL) OR (cliente_id = %s AND nombre = %s) LIMIT 1",
            (n_ship_to or None, cliente_id, ship_to),
        )
        fila_planta = cur.fetchone()
        if fila_planta:
            if ciudad and not fila_planta["ciudad"]:
                cur.execute("UPDATE planta SET ciudad = %s WHERE id = %s", (ciudad, fila_planta["id"]))
        else:
            cur.execute(
                "INSERT INTO planta (cliente_id, nombre, codigo_sap, ciudad, activo) VALUES (%s, %s, %s, %s, true)",
                (cliente_id, ship_to, n_ship_to or None, ciudad or None),
            )
            plantas_nuevas += 1

    return clientes_nuevos, plantas_nuevas


def _insertar_valores(cur, tipo: str, valores_crudos: list[str]) -> int:
    """Normaliza, agrupa por clave (colapsa duplicados de forma/mayúscula) y
    hace upsert -no inserta si ya existe una fila con esa clave (activa o
    fusionada, para no revivir algo que un admin ya homogenizó/desactivó)-."""
    por_clave: dict[str, list[str]] = {}
    for crudo in valores_crudos:
        normalizado = normalizar_texto_general(crudo)
        if not normalizado:
            continue
        clave = clave_normalizada(normalizado)
        por_clave.setdefault(clave, []).append(normalizado)

    insertados = 0
    for clave, variantes in por_clave.items():
        cur.execute("SELECT id FROM valor_lista WHERE tipo = %s AND valor_normalizado = %s", (tipo, clave))
        if cur.fetchone():
            continue
        representante = Counter(variantes).most_common(1)[0][0]
        cur.execute(
            "INSERT INTO valor_lista (tipo, valor, valor_normalizado, activo) VALUES (%s, %s, %s, true)",
            (tipo, representante, clave),
        )
        insertados += 1
    return insertados


def importar_especie_variedad(cur, wb) -> tuple[int, int]:
    ws = wb[HOJA_ESPECIE_VARIEDAD]
    filas = list(ws.iter_rows(values_only=True))[1:]

    especies = [_texto(f[0]) for f in filas if len(f) > 0 and _texto(f[0])]
    variedades = [_texto(f[2]) for f in filas if len(f) > 2 and _texto(f[2]) and _texto(f[2]) not in ("0",)]

    n_especies = _insertar_valores(cur, "especie", especies)
    n_variedades = _insertar_valores(cur, "variedad", variedades)
    return n_especies, n_variedades


def main() -> None:
    ruta = sys.argv[1] if len(sys.argv) > 1 else RUTA_POR_DEFECTO
    if not os.path.isfile(ruta):
        print(f"No se encontró el archivo: {ruta}")
        sys.exit(1)

    wb = load_workbook(ruta, data_only=True, read_only=True)
    with conexion() as conn, cursor_dict(conn) as cur:
        clientes_nuevos, plantas_nuevas = importar_sold_to_ship_to(cur, wb)
        especies_nuevas, variedades_nuevas = importar_especie_variedad(cur, wb)
    wb.close()

    print(f"Sold To nuevos: {clientes_nuevos}")
    print(f"Ship To nuevos: {plantas_nuevas}")
    print(f"Especie nuevas: {especies_nuevas}")
    print(f"Variedad nuevas: {variedades_nuevas}")
    print("Listo. Revisa 'Homogenizar' en el mantenedor de Listados para consolidar variantes de Variedad.")


if __name__ == "__main__":
    main()
