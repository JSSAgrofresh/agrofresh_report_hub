"""
Homogeniza Sold To, Ship To, Especie y Variedad de un Excel antes de subirlo
por Ingest, usando el catálogo oficial y el histórico ya resuelto a mano.

Tres usos, en el orden en que conviene hacerlos:

  1. Revisar el histórico antes de confiar en él
     python scripts/homogenizar_excel.py --catalogo CAT.xlsx --historico AYUDA.xlsx --conflictos

     Lista los valores crudos que fueron homogenizados a más de un destino.
     Mientras existan, el mismo texto puede resolverse de dos formas y no hay
     manera de que el script acierte siempre: hay que decidir cuál vale.

  2. Medir cuánto resuelve solo
     python scripts/homogenizar_excel.py --catalogo CAT.xlsx --historico AYUDA.xlsx --evaluar

     Reprocesa el histórico como si fuera nuevo y compara contra la columna ya
     homogenizada. Dice el porcentaje que se resuelve solo y, sobre todo, los
     casos en que el script se equivocó -que es lo que hay que mirar.

  3. Homogenizar un archivo nuevo
     python scripts/homogenizar_excel.py --catalogo CAT.xlsx --historico AYUDA.xlsx \\
            --archivo NUEVO.xlsx --salida NUEVO_HOMOGENIZADO.xlsx

     Escribe una copia con las columnas corregidas y una hoja "_pendientes"
     con lo que no se pudo resolver solo.

El catálogo es el Excel de valores oficiales: una columna por campo
(CROP, Variedad, SOLD TO2, SHIP TO2). El histórico es el Excel que trae el
par crudo/homogenizado (SOLD TO + SOLD TO2, SHIP TO + SHIP TO2).
"""
from __future__ import annotations

import argparse
import os
import re
import sys
from collections import Counter, defaultdict

import openpyxl

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

from app.homogenizador import Homogenizador, clave  # noqa: E402

# Campo -> (columna del catálogo, columna cruda, columna auxiliar donde se
# escribe el resultado). La auxiliar es opcional: si el archivo no la trae, se
# corrige la columna cruda en su lugar. Escribir aparte es lo preferible porque
# deja el dato original a la vista para poder revisar la decisión.
CAMPOS = {
    "Sold To": ("SOLD TO2", "SOLD TO", "SOLD TO2"),
    "Ship To": ("SHIP TO2", "SHIP TO", "SHIP TO2"),
    "Especie": ("CROP", "CROP", "CROP2"),
    "Variedad": ("Variedad", "Variedad", "Variedad2"),
}

# El orden importa: Ship To se resuelve después de Sold To porque lo usa como
# contexto, y Variedad después de Especie por la misma razón.
ORDEN = ("Sold To", "Ship To", "Especie", "Variedad")

# Restos de un buscar/reemplazar mal aplicado en el histórico. Son valores
# imposibles, no decisiones: usarlos como referencia enseñaría basura.
_CORRUPTO = re.compile(r"IFICOLIMARI|GESTIÓIFICOS|S A\.del", re.I)


def _hoja(ruta: str):
    wb = openpyxl.load_workbook(ruta, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    filas = list(ws.iter_rows(values_only=True))
    wb.close()
    if not filas:
        raise SystemExit(f"{ruta}: la primera hoja está vacía.")
    encabezado = [(str(c).strip() if c is not None else None) for c in filas[0]]
    indices = {n: i for i, n in enumerate(encabezado) if n}
    datos = [f for f in filas[1:] if any(c is not None for c in f)]
    return indices, datos


def _valor(fila, indices, columna) -> str:
    i = indices.get(columna)
    if i is None or i >= len(fila) or fila[i] is None:
        return ""
    return str(fila[i]).strip()


def _catalogo(ruta: str) -> dict[str, list[str]]:
    indices, datos = _hoja(ruta)
    salida = {}
    for campo, (col_cat, _, _) in CAMPOS.items():
        if col_cat not in indices:
            raise SystemExit(f"El catálogo no tiene la columna {col_cat!r}.")
        vistos = {}
        for fila in datos:
            v = _valor(fila, indices, col_cat)
            if v:
                vistos.setdefault(clave(v), v)
        salida[campo] = list(vistos.values())
    return salida


def _pares(ruta: str) -> dict[str, list[tuple[str, str]]]:
    """Pares (crudo, homogenizado) del histórico, sin las filas corruptas."""
    indices, datos = _hoja(ruta)
    salida: dict[str, list[tuple[str, str]]] = {}
    for campo, (_, col_crudo, col_ok) in CAMPOS.items():
        if not col_ok or col_ok == col_crudo or col_crudo not in indices or col_ok not in indices:
            continue
        pares = []
        for fila in datos:
            crudo = _valor(fila, indices, col_crudo)
            ok = _valor(fila, indices, col_ok)
            if crudo and ok and not _CORRUPTO.search(ok):
                pares.append((crudo, ok))
        salida[campo] = pares
    return salida


def _conflictos(pares: list[tuple[str, str]]) -> dict[str, Counter]:
    """Valores crudos que en el histórico apuntan a más de un oficial."""
    destinos: dict[str, Counter] = defaultdict(Counter)
    for crudo, ok in pares:
        destinos[clave(crudo)][ok] += 1
    return {k: v for k, v in destinos.items() if len(v) > 1}


def _homogenizador(campo: str, catalogo, pares, pares_ctx=None) -> Homogenizador:
    """El alias se siembra con el destino MAYORITARIO de cada valor crudo: si
    el histórico se contradice, gana lo que más veces decidió una persona.

    `pares_ctx` son tríos (contexto, crudo, oficial) para los campos que solo
    identifican dentro de otro -una sucursal dentro de su cliente-."""
    conteo: dict[str, Counter] = defaultdict(Counter)
    for crudo, ok in pares.get(campo, []):
        conteo[crudo][ok] += 1
    alias = {crudo: destinos.most_common(1)[0][0] for crudo, destinos in conteo.items()}

    conteo_ctx: dict[tuple[str, str], Counter] = defaultdict(Counter)
    for contexto, crudo, ok in (pares_ctx or []):
        conteo_ctx[(contexto, crudo)][ok] += 1
    alias_ctx = {k: d.most_common(1)[0][0] for k, d in conteo_ctx.items()}
    return Homogenizador(catalogo[campo], alias, alias_ctx)


def _pares_ship_por_cliente(ruta: str) -> list[tuple[str, str, str]]:
    """Tríos (cliente homogenizado, sucursal cruda, sucursal homogenizada).

    Una sucursal se llama igual en clientes distintos y se homogeniza distinto
    -"LONTUE" es "DOLE LONTUE" para DOLE y "LONTUE" para el resto-, así que
    por texto solo es irresoluble. El cliente es la única forma de decidir.
    """
    indices, datos = _hoja(ruta)
    if not {"SOLD TO2", "SHIP TO", "SHIP TO2"} <= set(indices):
        return []
    trios = []
    for fila in datos:
        cliente = _valor(fila, indices, "SOLD TO2")
        crudo = _valor(fila, indices, "SHIP TO")
        ok = _valor(fila, indices, "SHIP TO2")
        if cliente and crudo and ok and not _CORRUPTO.search(ok) and not _CORRUPTO.search(cliente):
            trios.append((cliente, crudo, ok))
    return trios


def _leer_decisiones(ruta: str) -> dict[str, dict[str, str]]:
    """Decisiones confirmadas a mano: {campo: {clave(crudo): oficial}}.

    Se lee la hoja "_decisiones" del archivo que produjo una corrida anterior,
    ya revisada. Una fila con DECISION vacía se ignora: significa que todavía
    no se decidió, no que deba borrarse el valor.
    """
    wb = openpyxl.load_workbook(ruta, read_only=True, data_only=True)
    if "_decisiones" not in wb.sheetnames:
        wb.close()
        raise SystemExit(f"{ruta}: no tiene la hoja '_decisiones'.")
    filas = list(wb["_decisiones"].iter_rows(values_only=True))
    wb.close()
    salida: dict[str, dict[str, str]] = defaultdict(dict)
    for fila in filas[1:]:
        if not fila or len(fila) < 5:
            continue
        campo = str(fila[0]).strip() if fila[0] else ""
        crudo = str(fila[1]).strip() if fila[1] else ""
        decision = str(fila[4]).strip() if fila[4] else ""
        if campo in CAMPOS and crudo and decision:
            salida[campo][clave(crudo)] = decision
    return salida


def cmd_conflictos(catalogo, pares) -> None:
    print("\nValores crudos homogenizados a más de un destino.")
    print("Mientras existan, el mismo texto puede resolverse de dos formas.\n")
    hay = False
    for campo, lista in pares.items():
        conf = _conflictos(lista)
        if not conf:
            print(f"  {campo}: sin conflictos.")
            continue
        hay = True
        afectadas = sum(sum(c.values()) for c in conf.values())
        print(f"\n=== {campo}: {len(conf)} valores ambiguos, {afectadas} filas ===")
        for k, destinos in sorted(conf.items(), key=lambda x: -sum(x[1].values())):
            print(f"\n  {k!r}")
            for dest, n in destinos.most_common():
                marca = "  <-- se usaría este" if dest == destinos.most_common(1)[0][0] else ""
                print(f"      {n:>5}x  {dest!r}{marca}")
    if hay:
        print("\nSe toma el destino mayoritario de cada uno. Corrige el histórico")
        print("si alguno debería resolverse al otro valor.")


def cmd_evaluar(catalogo, pares) -> None:
    print("\nSe reprocesa el histórico como si llegara nuevo y se compara")
    print("contra la columna ya homogenizada a mano.\n")
    for campo, lista in pares.items():
        if not lista:
            continue
        oficiales_k = {clave(o) for o in catalogo[campo]}
        # Un destino fuera del catálogo no es evaluable: no hay respuesta correcta.
        evaluables = [(c, o) for c, o in lista if clave(o) in oficiales_k]
        h = Homogenizador(catalogo[campo])
        reglas = Counter()
        auto = ok = 0
        errores = []
        for crudo, esperado in evaluables:
            r = h.resolver(crudo)
            reglas[r.regla] += 1
            if r.automatico:
                auto += 1
                if (r.valor or "").casefold() == esperado.casefold():
                    ok += 1
                else:
                    errores.append((crudo, r.valor, esperado, r.regla))
        total = len(evaluables)
        print(f"=== {campo}: {total} filas evaluables ===")
        print(f"  Resueltas solas : {auto:>5} ({auto / total * 100:5.1f}%)")
        print(f"    correctas     : {ok:>5} ({ok / max(auto, 1) * 100:5.1f}% de las automáticas)")
        print(f"    equivocadas   : {auto - ok:>5}")
        print(f"  A revisión      : {total - auto:>5} ({(total - auto) / total * 100:5.1f}%)")
        print(f"  Por regla: {dict(reglas.most_common())}")
        if errores:
            print("  Equivocaciones (hasta 10, agrupadas):")
            for (c, d, e, rg), n in Counter(errores).most_common(10):
                print(f"    {n:>4}x [{rg}] {c!r}\n           dio {d!r} / esperado {e!r}")
        print()


def cmd_homogenizar(catalogo, pares, ruta_historico: str, ruta_archivo: str, ruta_salida: str,
                    decisiones: dict[str, dict[str, str]] | None = None) -> None:
    indices, datos = _hoja(ruta_archivo)
    encabezado = [None] * (max(indices.values()) + 1)
    for nombre, i in indices.items():
        encabezado[i] = nombre

    ctx = _pares_ship_por_cliente(ruta_historico)
    hs = {
        campo: _homogenizador(campo, catalogo, pares, ctx if campo == "Ship To" else None)
        for campo in CAMPOS
    }
    # Las decisiones confirmadas entran como alias: pasan a resolverse solas.
    # Van después de construir el homogenizador para que ganen sobre lo que
    # hubiera aprendido del histórico.
    for campo, mapa in (decisiones or {}).items():
        for k, oficial in mapa.items():
            hs[campo].aprender(k, oficial)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Homogenizado"
    ws.append(encabezado)

    pendientes = [["Fila", "Campo", "Valor crudo", "Regla", "Sugerencias"]]
    resumen: dict[str, Counter] = {campo: Counter() for campo in CAMPOS}
    automaticas: Counter = Counter()

    for n_fila, fila in enumerate(datos, start=2):
        salida = list(fila) + [None] * (len(encabezado) - len(fila))
        # Sold To se resuelve primero porque Ship To lo necesita como contexto.
        cliente_resuelto = ""
        for campo in ORDEN:
            _, col_crudo, col_destino = CAMPOS[campo]
            # Se escribe en la auxiliar si existe; si no, sobre la cruda.
            i = indices.get(col_destino if col_destino in indices else col_crudo)
            if i is None:
                continue
            crudo = _valor(fila, indices, col_crudo)
            if not crudo:
                continue
            # Lo que una persona ya resolvió a mano manda: no se pisa, y además
            # sirve de contexto para los campos que se resuelven después.
            ya_resuelto = _valor(fila, indices, col_destino) if col_destino != col_crudo else ""
            if ya_resuelto:
                resumen[campo]["ya_resuelto"] += 1
                automaticas[campo] += 1
                if campo == "Sold To":
                    cliente_resuelto = ya_resuelto
                continue
            r = hs[campo].resolver(crudo, cliente_resuelto if campo == "Ship To" else None)
            resumen[campo][r.regla] += 1
            if r.automatico and r.valor:
                automaticas[campo] += 1
                salida[i] = r.valor
                if campo == "Sold To":
                    cliente_resuelto = r.valor
            else:
                pendientes.append([
                    n_fila, campo, crudo, r.regla,
                    " | ".join(f"{v} ({p:.0%})" for v, p in r.sugerencias) or "—",
                ])
        ws.append(salida)

    ws2 = wb.create_sheet("_pendientes")
    for f in pendientes:
        ws2.append(f)

    # Hoja de decisiones: una fila por valor DISTINTO, no por fila del Excel.
    # Lo pendiente son cientos de filas pero pocas decisiones -"Scarlett"
    # aparece 23 veces y se decide una sola vez-. La columna DECISION viene
    # con la mejor sugerencia ya escrita: revisar y corregir es más rápido
    # que escribir desde cero.
    agrupado: dict[tuple[str, str], dict] = {}
    for _fila, campo, crudo, regla, sugs in pendientes[1:]:
        d = agrupado.setdefault((campo, crudo), {"n": 0, "regla": regla, "sugs": sugs})
        d["n"] += 1
    ws3 = wb.create_sheet("_decisiones")
    ws3.append(["Campo", "Valor crudo", "Filas", "Regla", "DECISION (editar)", "Confianza", "Otras sugerencias"])
    for (campo, crudo), d in sorted(agrupado.items(), key=lambda x: (x[0][0], -x[1]["n"])):
        opciones = [t.strip() for t in (d["sugs"] or "").split("|") if t.strip() and t.strip() != "—"]
        mejor, confianza = "", ""
        if opciones:
            m = re.match(r"^(.*) \((\d+)%\)$", opciones[0])
            if m:
                mejor, confianza = m.group(1), f"{m.group(2)}%"
        ws3.append([campo, crudo, d["n"], d["regla"], mejor, confianza, " | ".join(opciones[1:])])
    ws3.freeze_panes = "A2"
    for col, ancho in zip("ABCDEFG", (12, 38, 8, 20, 38, 11, 60)):
        ws3.column_dimensions[col].width = ancho

    wb.save(ruta_salida)

    print(f"\nEscrito: {ruta_salida}")
    print(f"Filas: {len(datos)}\n")
    for campo, reglas in resumen.items():
        total = sum(reglas.values())
        if not total:
            continue
        auto = automaticas[campo]
        print(f"  {campo:<10} {auto:>5}/{total} resueltas ({auto / total * 100:5.1f}%)  {dict(reglas.most_common())}")
    print(f"\n  Pendientes: {len(pendientes) - 1} filas = {len(agrupado)} decisiones distintas")
    print("  Revisa la hoja '_decisiones': una fila por decisión, con la sugerencia ya escrita.")


def main() -> None:
    p = argparse.ArgumentParser(description=__doc__, formatter_class=argparse.RawDescriptionHelpFormatter)
    p.add_argument("--catalogo", required=True, help="Excel con los valores oficiales.")
    p.add_argument("--historico", required=True, help="Excel con el par crudo/homogenizado.")
    p.add_argument("--conflictos", action="store_true", help="Lista los valores ambiguos del histórico.")
    p.add_argument("--evaluar", action="store_true", help="Mide cuánto resuelve solo y qué falla.")
    p.add_argument("--archivo", help="Excel a homogenizar.")
    p.add_argument("--salida", help="Dónde escribir el resultado.")
    p.add_argument(
        "--decisiones",
        help="Excel de una corrida anterior con la hoja '_decisiones' ya revisada. "
             "Lo confirmado ahí se aplica y deja de aparecer como pendiente.",
    )
    args = p.parse_args()

    catalogo = _catalogo(args.catalogo)
    pares = _pares(args.historico)

    print("Catálogo oficial:", ", ".join(f"{c}={len(v)}" for c, v in catalogo.items()))
    print("Histórico:", ", ".join(f"{c}={len(v)} pares" for c, v in pares.items()))

    if args.conflictos:
        cmd_conflictos(catalogo, pares)
    if args.evaluar:
        cmd_evaluar(catalogo, pares)
    if args.archivo:
        if not args.salida:
            raise SystemExit("--archivo necesita --salida.")
        decisiones = _leer_decisiones(args.decisiones) if args.decisiones else None
        if decisiones:
            print("Decisiones confirmadas:", ", ".join(f"{c}={len(m)}" for c, m in decisiones.items()))
        cmd_homogenizar(catalogo, pares, args.historico, args.archivo, args.salida, decisiones)
    if not (args.conflictos or args.evaluar or args.archivo):
        p.print_help()


if __name__ == "__main__":
    main()
